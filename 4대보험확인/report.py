"""4대보험 확인 프로그램 — 엑셀 리포트 생성 모듈."""

from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ERROR_TYPE_COLORS = {
    "금액 불일치": "FFFF00",
    "공단에만 없음": "FFC000",
    "급여대장에만 없음": "ADD8E6",
    "전월 퇴사자 여전히 부과": "FF6666",
}

REPORT_HEADER = [
    "사원번호", "성명", "부서", "보험종류", "급여대장금액",
    "공단금액", "차액", "오류유형",
]


def build_report(merged_records, all_errors, 귀속년월, output_path):
    """'요약'과 '오류내역' 두 시트를 가진 엑셀 리포트를 output_path에 저장한다."""
    wb = Workbook()
    _write_summary_sheet(wb.active, merged_records, all_errors, 귀속년월)
    _write_error_detail_sheet(wb.create_sheet("오류내역"), all_errors)
    wb.save(output_path)


def _write_summary_sheet(ws, merged_records, all_errors, 귀속년월):
    ws.title = "요약"
    bold = Font(bold=True)

    ws.append(["4대보험료 확인 결과 요약"])
    ws["A1"].font = bold
    ws.append(["귀속년월", 귀속년월])
    ws.append(["전체 인원 수", len(merged_records)])
    ws.append(["전체 오류 건수", len(all_errors)])
    ws.append([])

    ws.append(["보험종류", "오류유형", "건수"])
    counts = Counter((e["보험종류"], e["오류유형"]) for e in all_errors)
    for (보험종류, 오류유형), count in sorted(counts.items()):
        ws.append([보험종류, 오류유형, count])


def _write_error_detail_sheet(ws, all_errors):
    ws.append(REPORT_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for error in all_errors:
        row = [
            error["사원번호"], error["성명"], error["부서"], error["보험종류"],
            error["급여대장금액"], error["공단금액"], error["차액"], error["오류유형"],
        ]
        ws.append(row)

        fill_color = ERROR_TYPE_COLORS.get(error["오류유형"])
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill
