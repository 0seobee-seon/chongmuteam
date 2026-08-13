import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook

from parsers import (
    normalize_birthdate,
    parse_employment_insurance,
    parse_health_insurance,
    parse_national_pension,
    parse_payroll,
    to_amount,
)


def test_normalize_birthdate_from_payroll_format():
    yymmdd, gender = normalize_birthdate("50.01.03(1)")
    assert yymmdd == "500103"
    assert gender == "1"


def test_normalize_birthdate_from_resident_number():
    yymmdd, gender = normalize_birthdate("900101-1234567")
    assert yymmdd == "900101"
    assert gender == "1"


def test_normalize_birthdate_from_employment_insurance_format():
    yymmdd, gender = normalize_birthdate("90-01-01")
    assert yymmdd == "900101"
    assert gender is None


def test_normalize_birthdate_none_input():
    yymmdd, gender = normalize_birthdate(None)
    assert yymmdd is None
    assert gender is None


def test_to_amount_from_int():
    assert to_amount(188010) == 188010


def test_to_amount_from_float():
    assert to_amount(204320.0) == 204320


def test_to_amount_from_comma_text():
    assert to_amount("48,280") == 48280


def test_to_amount_from_negative_comma_text():
    assert to_amount("-68,220") == -68220


def test_to_amount_none_is_none():
    assert to_amount(None) is None


def test_to_amount_blank_string_is_none():
    assert to_amount("  ") is None


def _make_payroll_workbook(path):
    """실제 급여대장 구조(직원 1명당 4행, 페이지 헤더 반복)를 흉내낸 합성 파일 생성.

    아래 값은 모두 가상의 테스트용 데이터이며 실제 직원 정보가 아니다.
    """
    wb = Workbook()
    ws = wb.active

    def write_row(row_idx, values):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 1페이지 헤더
    write_row(1, [None] * 8 + ["2026년 07월분  급여대장"])
    write_row(3, ["(주)테스트", None, None, "정렬 :", "코드순", None, None, None,
                  "[귀속:2026년 07월] [지급:2026년 07월 24일]"])
    write_row(6, ["사원번호", "성  명", None, "근무일수", "근무시간", "기본급",
                  "연장휴일근로수당", "야간수당", "전월소급", "감리수당", "출장수당",
                  "국민연금", "건강보험", "고용보험", "장기요양보험료"])
    write_row(7, ["입 사 일", "생년월일(성별)"])
    write_row(8, ["퇴 사 일", "종사업무", "부서", "직급"])
    write_row(9, ["임금계산 기초사항"])

    # 가상 직원 1: 국민연금/건강보험/고용보험/장기요양보험 모두 정상 값
    write_row(11, ["9001", "김테스트", None, "22일", "176시간", 39800000, None, None,
                   None, None, None, 500000, 1430810, 120000, 188010])
    write_row(12, ["1975-02-24", "90.01.01(1)"])
    write_row(13, [None, "사무업무", "총무팀", "대리"])
    write_row(14, ["기본 시간급, 일급, 월급 등"])

    # 가상 직원 2: 국민연금·고용보험 칸이 공란(가입 제외 대상 시나리오)
    write_row(15, ["9002", "이가상", None, "22일", "176시간", 19200000, None, None,
                   None, None, None, None, 690240, None, 90700])
    write_row(16, ["1992-02-13", "95.05.05(2)"])
    write_row(17, [None, "사무업무", "총무팀", "과장"])
    write_row(18, ["기본 시간급, 일급, 월급 등"])

    # 가상 직원 3: 동명이인 구분용으로 성명 뒤에 "(생년 2자리')" 접미사가 붙은 실제 표기 방식
    write_row(19, ["9003", "박동명(90')", None, "22일", "176시간", 20000000, None, None,
                   None, None, None, 300000, 500000, 40000, 60000])
    write_row(20, ["2020-03-01", "90.03.01(1)"])
    write_row(21, [None, "사무업무", "총무팀", "사원"])
    write_row(22, ["기본 시간급, 일급, 월급 등"])

    wb.save(path)


def test_parse_payroll_extracts_귀속년월(tmp_path):
    path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(str(path))

    귀속년월, records = parse_payroll(str(path))

    assert 귀속년월 == "202607"


def test_parse_payroll_extracts_all_employees(tmp_path):
    path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(str(path))

    _, records = parse_payroll(str(path))

    assert len(records) == 3
    assert records[0]["사원번호"] == "9001"
    assert records[0]["성명"] == "김테스트"
    assert records[0]["생년월일6자리"] == "900101"
    assert records[0]["성별코드"] == "1"
    assert records[0]["부서"] == "총무팀"
    assert records[0]["국민연금_급여대장"] == 500000
    assert records[0]["건강보험_급여대장"] == 1430810
    assert records[0]["고용보험_급여대장"] == 120000
    assert records[0]["장기요양보험_급여대장"] == 188010


def test_parse_payroll_blank_cells_mean_not_enrolled(tmp_path):
    path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(str(path))

    _, records = parse_payroll(str(path))

    이가상 = records[1]
    assert 이가상["국민연금_급여대장"] is None
    assert 이가상["고용보험_급여대장"] is None
    assert 이가상["건강보험_급여대장"] == 690240


def test_parse_payroll_strips_duplicate_name_birth_year_suffix(tmp_path):
    # 급여대장은 동명이인을 구분하려고 성명 뒤에 "(생년 2자리')"를 실제로 붙여서 저장한다
    # (예: "김성기(54')"). 공단 자료들은 이 접미사 없이 순수 성명만 있으므로, 매칭이
    # 되려면 파싱 단계에서 이 접미사를 제거해야 한다.
    path = tmp_path / "payroll.xlsx"
    _make_payroll_workbook(str(path))

    _, records = parse_payroll(str(path))

    박동명 = records[2]
    assert 박동명["성명"] == "박동명"


def _make_health_insurance_workbook(path):
    """실제 '보험료 고지(산출) 내역서' 헤더 구조를 흉내낸 합성 파일.

    아래 값은 모두 가상의 테스트용 데이터이며 실제 직원 정보가 아니다.
    """
    wb = Workbook()
    ws = wb.active
    header = [
        "고지년월", "사업장관리번호", "단위사업장(단위기관)", "고지차수", "회계",
        "증번호", "성명", "주민등록번호", "감면사유", "직종", "등급", "보수월액",
        "산출보험료", "정산사유", "시작월", "종료월", "정산금액", "고지금액",
        "연말정산", "취득일", "상실일", "요양산출보험료", "요양정산사유코드",
        "요양시작월", "요양종료월", "요양정산보험료", "요양고지보험료",
    ]
    ws.append(header)
    ws.append([
        202607, 10000000000, "000", 1, "00", 90000000001, "김테스트",
        "900101-1234567", "00", "00", None, 5683666.0, 204320.0, None, None,
        None, 0.0, 204320.0, 0.0, 20060627, 99991231, 26840.0, None, None,
        None, 0.0, 26840.0,
    ])
    ws.append([
        202607, 10000000000, "000", 1, "00", 90000000002, "이가상",
        "950505-2345678", "00", "00", None, 2922000.0, 105040.0, None, None,
        None, 0.0, 105040.0, 0.0, 20240401, 20260630, 13800.0, None, None,
        None, 0.0, 13800.0,
    ])
    wb.save(path)


def test_parse_health_insurance_extracts_records(tmp_path):
    path = tmp_path / "health.xlsx"
    _make_health_insurance_workbook(str(path))

    records = parse_health_insurance(str(path))

    assert len(records) == 2
    assert records[0]["성명"] == "김테스트"
    assert records[0]["생년월일6자리"] == "900101"
    assert records[0]["성별코드"] == "1"
    assert records[0]["건강보험_공단"] == 204320
    assert records[0]["장기요양보험_공단"] == 26840
    assert records[0]["상실일"] == 99991231


def test_parse_health_insurance_second_row_has_termination_date(tmp_path):
    path = tmp_path / "health.xlsx"
    _make_health_insurance_workbook(str(path))

    records = parse_health_insurance(str(path))

    assert records[1]["상실일"] == 20260630


def _make_national_pension_workbook(path):
    """실제 국민연금공단 '가입자내역' 헤더 구조(2줄 헤더, 데이터는 5행부터)를 흉내낸 합성 파일.

    아래 값은 모두 가상의 테스트용 데이터이며 실제 직원 정보가 아니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["가입자내역"])
    ws.append([])
    ws.append(["순번", "연금보험료 대상자", None, "보험료지원금 공제 전", None, None, None,
               "보험료지원금 공제 후(대상자만 표기)"])
    ws.append([None, "성명", "주민등록번호", "기준소득월액(원)", "연금보험료(원)",
               "근로자기여금(원)", "사용자부담금(원)", "근로자기여금(원)", "사용자부담금(원)"])
    # 일반 직원 (지원금 공제 대상 아님 -> 공제 전 금액이 실제 부과액)
    ws.append([1, "김테스트", "900101-1234567", 2609000.0, 247840.0, 123920.0, 123920.0, None, None])
    # 두루누리 지원 대상 직원 (공제 후 금액이 실제 부과액)
    ws.append([2, "이가상", "950505-2345678", 2200000.0, 209000.0, 104500.0, 104500.0, 52250.0, 104500.0])
    wb.save(path)


def test_parse_national_pension_uses_공제전_금액_when_no_지원금(tmp_path):
    path = tmp_path / "pension.xlsx"
    _make_national_pension_workbook(str(path))

    records = parse_national_pension(str(path))

    assert records[0]["성명"] == "김테스트"
    assert records[0]["생년월일6자리"] == "900101"
    assert records[0]["국민연금_공단"] == 123920


def test_parse_national_pension_uses_공제후_금액_when_지원금_있음(tmp_path):
    path = tmp_path / "pension.xlsx"
    _make_national_pension_workbook(str(path))

    records = parse_national_pension(str(path))

    assert records[1]["성명"] == "이가상"
    assert records[1]["국민연금_공단"] == 52250


def _make_employment_insurance_workbook(path):
    """실제 근로복지공단 '당월보험료부과내역조회(고용)' 구조(헤더 2행, 데이터는 3행부터)를
    흉내낸 합성 파일.

    열 그룹 구조: 10~13=산정보험료(해당월①), 14~17=재산정보험료(해당년도②),
    18=정산보수총액, 19~22=정산보험료(③), 23~26=보험료합계(①+②+③). 각 4개 그룹은
    [근로자실업급여보험료, 사업주실업급여보험료, 사업주고용안정직능보험료, 합계] 순서.
    급여대장과 비교할 값은 23번째 열(보험료합계 그룹의 근로자실업급여보험료, 직원 본인
    부담분 최종 금액)이지 26번째 열(사업주 부담분까지 합친 그룹 전체 합계)이 아니다.
    실제 샘플에서 이 합계 열들은 쉼표 포함 텍스트로 저장되어 있었다.

    아래 값은 모두 가상의 테스트용 데이터이며 실제 직원 정보가 아니다.
    """
    wb = Workbook()
    ws = wb.active
    row1 = ["순번", "근로자구분", "근로자명", "생년월일", "근로자원부번호", "고용일",
            "고용종료일", "휴직자월평", "월평균보수금액"] + [None] * 17
    row2 = [None] * 26
    ws.append(row1)
    ws.append(row2)

    def make_row(순번, 이름, 생년월일, 고용일, 고용종료일, 근로자부담_최종, 그랜드합계):
        row = [순번, "일반", 이름, 생년월일, "000099999999", 고용일, 고용종료일, None, 5000000.0]
        row += [None] * 4  # 10~13 산정보험료 그룹(이번 테스트에서는 값 미검증)
        row += [None] * 4  # 14~17 재산정보험료 그룹
        row += [None]       # 18 정산보수총액
        row += [None] * 4  # 19~22 정산보험료(③) 그룹
        row += [근로자부담_최종, None, None, 그랜드합계]  # 23~26 보험료합계 그룹
        return row

    ws.append(make_row(1, "김테스트", "90-01-01", "2006-06-27", None, 50000.0, "108,000"))
    ws.append(make_row(2, "이가상", "95-05-05", "2024-04-01", "2026-06-30", "-10,000", "-22,000"))
    wb.save(path)


def test_parse_employment_insurance_extracts_employee_only_final_total(tmp_path):
    path = tmp_path / "employment.xlsx"
    _make_employment_insurance_workbook(str(path))

    records = parse_employment_insurance(str(path))

    assert records[0]["성명"] == "김테스트"
    assert records[0]["생년월일6자리"] == "900101"
    # 23번째 열(근로자 본인 부담분)을 읽어야 한다 — 26번째 열(사업주 부담분 포함 그랜드 합계)이 아님
    assert records[0]["고용보험_공단"] == 50000
    assert records[0]["고용종료일"] is None


def test_parse_employment_insurance_handles_negative_comma_amount(tmp_path):
    path = tmp_path / "employment.xlsx"
    _make_employment_insurance_workbook(str(path))

    records = parse_employment_insurance(str(path))

    assert records[1]["고용보험_공단"] == -10000
    assert records[1]["고용종료일"] == "2026-06-30"
