import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook

from report import build_report


def test_build_report_creates_요약_and_오류내역_sheets(tmp_path):
    merged_records = [{"사원번호": "0001", "성명": "김테스트"}]
    errors = [{
        "사원번호": "0001", "성명": "김테스트", "부서": "총무팀",
        "보험종류": "건강보험", "급여대장금액": 200000, "공단금액": 190000,
        "차액": 10000, "오류유형": "금액 불일치",
    }]
    output_path = tmp_path / "report.xlsx"

    build_report(merged_records, errors, "202607", str(output_path))

    wb = load_workbook(str(output_path))
    assert "요약" in wb.sheetnames
    assert "오류내역" in wb.sheetnames


def test_build_report_요약_sheet_has_error_counts(tmp_path):
    merged_records = [{"사원번호": "0001", "성명": "김테스트"}]
    errors = [
        {"사원번호": "0001", "성명": "김테스트", "부서": "총무팀", "보험종류": "건강보험",
         "급여대장금액": 200000, "공단금액": 190000, "차액": 10000, "오류유형": "금액 불일치"},
        {"사원번호": "0002", "성명": "김철수", "부서": "총무팀", "보험종류": "국민연금",
         "급여대장금액": None, "공단금액": 100000, "차액": None, "오류유형": "급여대장에만 없음"},
    ]
    output_path = tmp_path / "report.xlsx"

    build_report(merged_records, errors, "202607", str(output_path))

    wb = load_workbook(str(output_path))
    summary_values = [cell.value for row in wb["요약"].iter_rows() for cell in row]
    assert "202607" in [str(v) for v in summary_values]
    assert 1 in summary_values  # 전체 인원 수


def test_build_report_오류내역_sheet_lists_each_error_row(tmp_path):
    merged_records = [{"사원번호": "0001", "성명": "김테스트"}]
    errors = [{
        "사원번호": "0001", "성명": "김테스트", "부서": "총무팀",
        "보험종류": "건강보험", "급여대장금액": 200000, "공단금액": 190000,
        "차액": 10000, "오류유형": "금액 불일치",
    }]
    output_path = tmp_path / "report.xlsx"

    build_report(merged_records, errors, "202607", str(output_path))

    wb = load_workbook(str(output_path))
    ws = wb["오류내역"]
    header = [cell.value for cell in ws[1]]
    data_row = [cell.value for cell in ws[2]]

    assert header[:5] == ["사원번호", "성명", "부서", "보험종류", "급여대장금액"]
    assert data_row[0] == "0001"
    assert data_row[3] == "건강보험"


def test_build_report_applies_fill_color_per_error_type(tmp_path):
    errors = [{
        "사원번호": "0001", "성명": "김테스트", "부서": "총무팀",
        "보험종류": "건강보험", "급여대장금액": 200000, "공단금액": 190000,
        "차액": 10000, "오류유형": "금액 불일치",
    }]
    output_path = tmp_path / "report.xlsx"

    build_report([{"사원번호": "0001", "성명": "김테스트"}], errors, "202607", str(output_path))

    wb = load_workbook(str(output_path))
    ws = wb["오류내역"]
    fill = ws.cell(row=2, column=8).fill  # 오류유형 열 (헤더: 사원번호..오류유형 8개 열 중 8번째)
    assert fill.fgColor.rgb != "00000000"
