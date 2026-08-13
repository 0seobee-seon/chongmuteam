"""4대보험 확인 프로그램 — 파일 파서 모듈.

급여대장, 건강보험공단, 국민연금공단, 근로복지공단(고용보험) 4개
엑셀 파일을 각각 직원별 레코드 리스트로 변환한다.
"""

import re

from openpyxl import load_workbook

EMPLOYEE_NUMBER_PATTERN = re.compile(r"^\d{4}$")
NAME_BIRTH_YEAR_SUFFIX_PATTERN = re.compile(r"\(\d+'?\)\s*$")


def strip_duplicate_name_suffix(성명):
    """급여대장은 동명이인을 구분하려고 성명 뒤에 "(생년 2자리')"를 붙여 저장한다
    (예: '김성기(54')'). 공단 자료에는 이 접미사가 없어 매칭이 깨지므로 제거한다.
    """
    return NAME_BIRTH_YEAR_SUFFIX_PATTERN.sub("", 성명).strip()


def normalize_birthdate(raw):
    """다양한 형식의 생년월일 문자열에서 YYMMDD 6자리(+가능하면 성별코드)를 뽑는다.

    지원 형식 예: '50.01.03(1)'(급여대장), '900101-1234567'(주민등록번호),
    '90-01-01'(고용보험, 성별코드 없음).
    """
    if not raw:
        return None, None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) < 6:
        return None, None
    yymmdd = digits[:6]
    gender = digits[6] if len(digits) >= 7 else None
    return yymmdd, gender


def to_amount(value):
    """엑셀 셀 값을 정수 금액으로 변환한다. 공란/None은 '가입대상 아님'을 뜻하므로 None을 반환."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if cleaned == "":
            return None
        return int(float(cleaned))
    return int(value)


def parse_payroll(path):
    """급여대장(회사 내부 급여관리 프로그램 출력본)을 파싱한다.

    인쇄용 리포트 형식이라 직원 1명당 4행이 한 블록이고, 페이지 헤더가
    반복된다. A열이 4자리 숫자면 블록 시작으로 인식해 4행씩 읽는다.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    귀속년월 = None
    records = []

    row = 1
    max_row = ws.max_row
    max_col = ws.max_column
    while row <= max_row:
        a_val = ws.cell(row=row, column=1).value

        if 귀속년월 is None:
            for col in range(1, max_col + 1):
                cell_val = ws.cell(row=row, column=col).value
                if isinstance(cell_val, str) and "귀속:" in cell_val:
                    match = re.search(r"귀속:(\d{4})년\s*(\d{2})월", cell_val)
                    if match:
                        귀속년월 = f"{match.group(1)}{match.group(2)}"
                    break

        a_str = str(a_val).strip() if a_val is not None else ""
        if EMPLOYEE_NUMBER_PATTERN.match(a_str) and row + 2 <= max_row:
            생년월일_원본 = ws.cell(row=row + 1, column=2).value
            생년월일6자리, 성별코드 = normalize_birthdate(생년월일_원본)

            성명_원본 = str(ws.cell(row=row, column=2).value or "").strip()

            records.append({
                "사원번호": a_str,
                "성명": strip_duplicate_name_suffix(성명_원본),
                "생년월일6자리": 생년월일6자리,
                "성별코드": 성별코드,
                "부서": str(ws.cell(row=row + 2, column=3).value or "").strip(),
                "퇴사일": ws.cell(row=row + 2, column=1).value,
                "국민연금_급여대장": to_amount(ws.cell(row=row, column=12).value),
                "건강보험_급여대장": to_amount(ws.cell(row=row, column=13).value),
                "고용보험_급여대장": to_amount(ws.cell(row=row, column=14).value),
                "장기요양보험_급여대장": to_amount(ws.cell(row=row, column=15).value),
            })
            row += 4
        else:
            row += 1

    return 귀속년월, records


def parse_health_insurance(path):
    """건강보험공단 '보험료 고지(산출) 내역서' 엑셀을 파싱한다. 1행 = 1직원."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        성명 = row[col["성명"]]
        if not 성명:
            continue
        생년월일6자리, 성별코드 = normalize_birthdate(row[col["주민등록번호"]])
        records.append({
            "성명": str(성명).strip(),
            "생년월일6자리": 생년월일6자리,
            "성별코드": 성별코드,
            "건강보험_공단": to_amount(row[col["고지금액"]]),
            "장기요양보험_공단": to_amount(row[col["요양고지보험료"]]),
            "상실일": row[col["상실일"]],
        })
    return records


def parse_national_pension(path):
    """국민연금공단 '가입자내역' 엑셀을 파싱한다.

    헤더가 3~4행에 걸쳐 있고 데이터는 5행부터 시작한다. 열 이름이 중복되므로
    (근로자기여금(원)이 두 번 나옴) 이름이 아닌 실제 열 위치로 읽는다:
    B=성명, C=주민등록번호, F=근로자기여금(공제 전), H=근로자기여금(공제 후).
    '공제 후'는 두루누리 사회보험료 지원 대상자만 값이 있고, 있으면 그 값이
    실제 부과액이다.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row_idx in range(5, ws.max_row + 1):
        성명 = ws.cell(row=row_idx, column=2).value
        if not 성명:
            continue
        생년월일6자리, 성별코드 = normalize_birthdate(ws.cell(row=row_idx, column=3).value)
        공제전 = ws.cell(row=row_idx, column=6).value
        공제후 = ws.cell(row=row_idx, column=8).value
        실제금액 = 공제후 if 공제후 not in (None, "") else 공제전

        records.append({
            "성명": str(성명).strip(),
            "생년월일6자리": 생년월일6자리,
            "성별코드": 성별코드,
            "국민연금_공단": to_amount(실제금액),
        })
    return records


def parse_employment_insurance(path):
    """근로복지공단 '당월보험료부과내역조회(고용)' 엑셀을 파싱한다.

    헤더가 1~2행이고 데이터는 3행부터 시작한다. 열 그룹 구조:
    10~13=산정보험료(해당월①), 14~17=재산정보험료(해당년도②), 18=정산보수총액,
    19~22=정산보험료(③), 23~26=보험료합계(①+②+③). 각 4개 그룹은
    [근로자실업급여보험료, 사업주실업급여보험료, 사업주고용안정직능보험료, 합계]
    순서의 서브컬럼을 가진다.

    급여대장에 입력되는 고용보험료는 직원 본인 부담분(근로자실업급여보험료)뿐이고
    사업주 부담분(실업급여+고용안정직능보험료)은 포함하지 않는다. 따라서 비교
    대상 금액은 26번째 열(그룹 전체 합계, 사업주 부담분 포함)이 아니라
    **23번째 열**(보험료합계 그룹의 근로자실업급여보험료 서브컬럼 — 산정·재산정·정산
    소급분이 모두 반영된 직원 본인 부담분 최종 금액)이다. 실제 샘플 데이터 한 건으로
    급여대장 값과 대조해 확인했다(급여대장 금액이 23번째 열과 일치하고, 26번째 열과는
    불일치함을 확인).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row_idx in range(3, ws.max_row + 1):
        성명 = ws.cell(row=row_idx, column=3).value
        if not 성명:
            continue
        생년월일6자리, _성별코드 = normalize_birthdate(ws.cell(row=row_idx, column=4).value)

        records.append({
            "성명": str(성명).strip(),
            "생년월일6자리": 생년월일6자리,
            "고용보험_공단": to_amount(ws.cell(row=row_idx, column=23).value),
            "고용종료일": ws.cell(row=row_idx, column=7).value,
        })
    return records
