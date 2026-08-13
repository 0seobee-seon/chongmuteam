"""일용직 근로내용확인신고서 생성 — 핵심 로직 (UI 없음).

급여 프로그램에서 뽑은 일용직 자료를 읽어 근로복지공단 전자신고용 서식을
현장별로 채운다. 신고는 사업장(현장) 단위이므로 현장이 섞인 자료를 넣어도
현장별로 파일이 나뉘어 생성된다.

지원하는 입력 형식
  A) [일용직 급여 지급명세서]  1인 2행. 상단행 1~15일 / 하단행 16~31일에 '●'.
                               일자별 출근 자료가 있어 그대로 신고서를 만들 수 있다.
  B) [현장별 일용직 급여내역]  1인 1행. 근무일수 합계만 있고 일자별 자료가 없다.
                               10~40열(1~31일)을 채울 수 없으므로 형식 A를 함께
                               넣어 주민번호로 보완해야 한다.
"""

import os
import re
import unicodedata

import openpyxl

# 전자신고용 '서식' 시트 열 번호 (1-based)
COL_INSURANCE = 1
COL_NAME = 2
COL_JUMIN = 3
COL_NATION = 4
COL_STAY = 5
COL_TEL_AREA = 6
COL_TEL_EXCH = 7
COL_TEL_LAST = 8
COL_JIKJONG = 9
COL_DAY1 = 10           # 10~40 = 1일~31일
COL_WORK_DAYS = 41
COL_AVG_HOURS = 42
COL_PAY_BASE_DAYS = 43
COL_TAXABLE = 44        # 보수총액(과세소득)
COL_WAGE_TOTAL = 45     # 임금총액
COL_RESIGN = 46
COL_NTS_YN = 49
COL_PAY_MONTH = 50
COL_NTS_TOTAL = 51
COL_NTS_EXEMPT = 52
COL_NTS_TAX = 53
COL_NTS_LOCAL_TAX = 54

SHEET_NAME = "서식"
FIRST_DATA_ROW = 2

# 보험구분 3·5(고용보험 포함)일 때만 이직사유코드가 필수값이다.
INSURANCE_OPTIONS = [("1", "산재보험"), ("3", "고용보험"), ("5", "산재 및 고용보험")]
RESIGN_OPTIONS = [
    ("1", "회사사정 (폐업·공사중단·공사종료·계약만료)"),
    ("2", "부득이한 개인사정 (질병·부상·출산)"),
    ("3", "기타 개인사정 (전직·자영업)"),
]

# 회사 표준값 — 매월 같은 값을 쓰므로 기본값으로 둔다.
DEFAULT_PARAMS = {
    "jikjong": "706",
    "insurance": "3",
    "resign": "1",
    "avg_hours": 8,
    "nts": False,       # 국세청 일용근로소득 신고여부. False면 49~54열을 비운다.
}

_TOTAL_MARKERS = ("합계", "총인원", "소계")
_FORMAT_A = "A"
_FORMAT_B = "B"


class LedgerFormatError(Exception):
    """대장 형식을 알아볼 수 없을 때."""


# ---------------------------------------------------------------------------
# 셀 값 정리
# ---------------------------------------------------------------------------

def _norm(value):
    """비교용으로 공백과 전각 문자를 정리한 문자열."""
    if value is None:
        return ""
    return re.sub(r"\s+", "", unicodedata.normalize("NFKC", str(value)))


def _is_total_row(value):
    text = _norm(value)
    return any(text.startswith(marker) for marker in _TOTAL_MARKERS)


def _digits(value):
    return re.sub(r"[^0-9]", "", str(value)) if value is not None else ""


def _to_int(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = re.sub(r"[^0-9\-]", "", str(value))
    return int(text) if text not in ("", "-") else 0


def _has_mark(row, index):
    return index < len(row) and row[index] not in (None, "", 0)


def normalize_month(value):
    """'2026-07' / '2026년07월' / date → '202607'. 알 수 없으면 None."""
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month"):
        return f"{value.year}{value.month:02d}"
    match = re.search(r"(\d{4})\D*(\d{1,2})", str(value))
    return f"{match.group(1)}{int(match.group(2)):02d}" if match else None


def is_foreigner(jumin, nat_flag=None):
    """주민번호 7번째 자리가 5~9면 외국인. 대장의 '내/외' 표기도 함께 본다."""
    by_jumin = len(jumin) >= 7 and jumin[6] in "56789"
    by_flag = _norm(nat_flag) == "외" if nat_flag is not None else False
    return by_jumin or by_flag


def split_phone(value):
    """전화번호를 (지역번호, 국번, 뒷번호)로 나눈다. 판단 불가면 (None, None, None)."""
    if not value:
        return None, None, None
    digits = _digits(value)
    if len(digits) == 11:
        return digits[:3], digits[3:7], digits[7:]
    if len(digits) == 10:
        return digits[:3], digits[3:6], digits[6:]
    return None, None, None


# ---------------------------------------------------------------------------
# 대장 읽기
# ---------------------------------------------------------------------------

def _load_rows(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    return list(sheet.iter_rows(values_only=True))


def detect_format(rows):
    for row in rows[:12]:
        joined = "".join(_norm(v) for v in row if v is not None)
        if "현장코드" in joined and "귀속월" in joined:
            return _FORMAT_B
        if "주민등록번호" in joined and "전화번호" in joined and "입사일" in joined:
            return _FORMAT_A
    raise LedgerFormatError(
        "일용직 자료 형식을 알아볼 수 없습니다. "
        "[일용직 급여 지급명세서] 또는 [현장별 일용직 급여내역] 파일을 넣어주세요."
    )


def _parse_format_b(rows, source):
    """[현장별 일용직 급여내역] — 헤더 이름으로 열 위치를 찾는다."""
    header_index = next(
        (i for i, row in enumerate(rows[:12])
         if any(_norm(v) == "현장코드" for v in row if v is not None)),
        None,
    )
    if header_index is None:
        raise LedgerFormatError(f"'현장코드' 헤더를 찾지 못했습니다: {source}")

    header = {}
    for index, value in enumerate(rows[header_index]):
        name = _norm(value)
        if name:
            header.setdefault(name, index)

    def col(*names):
        return next((header[n] for n in names if n in header), None)

    c_site = col("현장명")
    c_name = col("사원명", "성명")
    c_flag = col("주민(외국인)번호", "주민외국인번호")
    c_jumin = c_flag + 1 if c_flag is not None else None
    c_paymonth = col("지급월")
    c_days = col("근무일", "근무일수")
    c_hours = col("근무시간")
    c_pay = col("지급액")
    c_exempt = col("비과세")
    c_tax = col("소득세")
    c_local = col("지방소득세")
    c_wage = col("임금총액")
    c_tel = col("전화번호", "휴대폰", "연락처")

    missing = [name for name, index in
               (("현장명", c_site), ("사원명", c_name), ("근무일", c_days), ("지급액", c_pay))
               if index is None]
    if missing:
        raise LedgerFormatError(f"헤더에 {', '.join(missing)} 항목이 없습니다: {source}")

    people = []
    for row in rows[header_index + 1:]:
        if row is None or not any(v is not None for v in row):
            continue
        if _is_total_row(row[0]):
            break

        name = row[c_name] if c_name < len(row) else None
        jumin = _digits(row[c_jumin]) if c_jumin is not None and c_jumin < len(row) else ""
        if not name or len(jumin) != 13:
            continue

        pay = _to_int(row[c_pay])
        exempt = _to_int(row[c_exempt]) if c_exempt is not None else 0
        people.append({
            "site": str(row[c_site]).strip() if c_site is not None and row[c_site] else "",
            "name": str(name).strip(),
            "jumin": jumin,
            "nat_flag": row[c_flag] if c_flag is not None and c_flag < len(row) else None,
            "tel": row[c_tel] if c_tel is not None and c_tel < len(row) else None,
            "days": _to_int(row[c_days]),
            "hours": _to_int(row[c_hours]) if c_hours is not None else 0,
            "pay": pay,
            "exempt": exempt,
            "taxable": pay - exempt,
            "wage": _to_int(row[c_wage]) if c_wage is not None else pay,
            "tax": _to_int(row[c_tax]) if c_tax is not None else 0,
            "local_tax": _to_int(row[c_local]) if c_local is not None else 0,
            "paymonth": normalize_month(row[c_paymonth]) if c_paymonth is not None else None,
            "day_marks": None,
            "source": source,
        })
    return people


def _parse_format_a(rows, source):
    """[일용직 급여 지급명세서] — 1인 2행, 일자별 '●' 출근표."""
    site = ""
    paymonth = None
    for row in rows[:8]:
        for index, value in enumerate(row):
            if value is None:
                continue
            name = _norm(value)
            if "현장명" in name and not site:
                site = next((str(row[k]).strip()
                             for k in range(index + 1, min(index + 4, len(row))) if row[k]), "")
            if name.startswith("귀속") and paymonth is None:
                paymonth = next((normalize_month(row[k])
                                 for k in range(index + 1, min(index + 4, len(row)))
                                 if normalize_month(row[k])), None)

    people = []
    i = 8   # 데이터는 9행(인덱스 8)부터 시작한다.
    while i + 1 < len(rows):
        top, bottom = rows[i], rows[i + 1]
        if _is_total_row(top[0] if top else None):
            break

        jumin = _digits(top[2]) if len(top) > 2 else ""
        if len(jumin) != 13:
            i += 1
            continue

        marks = [1 if _has_mark(top, c) else None for c in range(5, 20)]      # 1~15일
        marks += [1 if _has_mark(bottom, c) else None for c in range(5, 20)]  # 16~30일
        marks.append(1 if _has_mark(bottom, 20) else None)                    # 31일

        pay = _to_int(top[24]) if len(top) > 24 else 0
        people.append({
            "site": site,
            "name": str(bottom[0]).strip() if bottom[0] else "",
            "jumin": jumin,
            "nat_flag": None,
            "tel": top[3] if len(top) > 3 else None,
            "days": _to_int(top[21]) if len(top) > 21 else sum(1 for m in marks if m),
            "hours": _to_int(top[22]) if len(top) > 22 else 0,
            "pay": pay,
            "exempt": 0,
            "taxable": pay,
            "wage": pay,
            "tax": _to_int(top[26]) if len(top) > 26 else 0,
            "local_tax": _to_int(bottom[26]) if len(bottom) > 26 else 0,
            "paymonth": paymonth,
            "day_marks": marks,
            "source": source,
        })
        i += 2
    return people


def parse_ledger(path):
    """대장 한 개를 읽어 (형식, 인원목록)을 돌려준다."""
    rows = _load_rows(path)
    fmt = detect_format(rows)
    source = os.path.basename(path)
    people = _parse_format_b(rows, source) if fmt == _FORMAT_B else _parse_format_a(rows, source)
    return fmt, people


def fill_missing_day_marks(people, grid_people):
    """일자별 자료가 없는 인원을 주민번호로 매칭해 보완한다. 보완한 인원 수를 돌려준다."""
    grid = {p["jumin"]: p["day_marks"] for p in grid_people if p["day_marks"]}
    filled = 0
    for person in people:
        if not person["day_marks"] and person["jumin"] in grid:
            person["day_marks"] = grid[person["jumin"]]
            filled += 1
    return filled


def group_by_site(people):
    """현장명으로 묶는다. 현장명이 비어 있으면 '현장미상'."""
    sites = {}
    for person in people:
        sites.setdefault(person["site"] or "현장미상", []).append(person)
    return sites


def dedupe_people(people):
    """같은 현장·같은 주민번호는 한 사람으로 합친다.

    같은 현장 자료를 [현장별 일용직 급여내역]과 [일용직 급여 지급명세서] 두 형식으로
    함께 넣어도 인원이 두 배로 잡히지 않게 한다. 일자별 출근표가 있는 쪽 값을 살린다.

    (인원목록, 합쳐진 건수)를 돌려준다.
    """
    merged = {}
    order = []
    duplicates = 0

    for person in people:
        key = (person["site"] or "현장미상", person["jumin"])
        existing = merged.get(key)
        if existing is None:
            merged[key] = person
            order.append(key)
            continue

        duplicates += 1
        if not existing["day_marks"] and person["day_marks"]:
            # 일자별 자료가 있는 쪽으로 교체하되, 먼저 읽은 순서는 유지한다.
            merged[key] = person

    return [merged[key] for key in order], duplicates


# ---------------------------------------------------------------------------
# 신고서 작성
# ---------------------------------------------------------------------------

def safe_filename(text):
    return re.sub(r'[\\/:*?"<>|]', "_", text).strip() or "현장미상"


def report_filename(site, paymonth):
    return f"근로내용확인신고_{safe_filename(site)}_{paymonth or '연월미상'}.xlsx"


def write_site_report(template_path, out_path, people, params):
    """한 현장의 신고서를 만든다. 확인이 필요한 항목을 dict로 돌려준다."""
    workbook = openpyxl.load_workbook(template_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise LedgerFormatError(
            f"신고서 양식에 '{SHEET_NAME}' 시트가 없습니다: {os.path.basename(template_path)}"
        )
    sheet = workbook[SHEET_NAME]

    insurance = str(params["insurance"])
    jikjong = str(params["jikjong"]).strip()
    resign = str(params["resign"])
    avg_hours = int(params["avg_hours"])
    report_nts = bool(params["nts"])

    site_paymonth = next((p["paymonth"] for p in people if p["paymonth"]), None)
    foreigners = []
    missing_days = []
    day_mismatch = []

    for offset, person in enumerate(people):
        row = FIRST_DATA_ROW + offset
        foreign = is_foreigner(person["jumin"], person["nat_flag"])

        sheet.cell(row, COL_INSURANCE, insurance)
        sheet.cell(row, COL_NAME, person["name"])
        sheet.cell(row, COL_JUMIN, person["jumin"])
        # 외국인은 국적·체류자격을 급여대장에서 알 수 없다. 셀을 아예 비워
        # 제출 전에 눈에 띄도록 두고, 대상자 명단을 따로 돌려준다.
        if not foreign:
            sheet.cell(row, COL_NATION, "100")

        area, exchange, last = split_phone(person["tel"])
        if area:
            sheet.cell(row, COL_TEL_AREA, area)
            sheet.cell(row, COL_TEL_EXCH, exchange)
            sheet.cell(row, COL_TEL_LAST, last)

        if jikjong:
            sheet.cell(row, COL_JIKJONG, jikjong)

        marks = person["day_marks"]
        if marks:
            for day_offset, mark in enumerate(marks):
                if mark:
                    sheet.cell(row, COL_DAY1 + day_offset, 1)
            if sum(1 for m in marks if m) != person["days"]:
                day_mismatch.append(person["name"])
        else:
            missing_days.append(person["name"])

        sheet.cell(row, COL_WORK_DAYS, person["days"])
        sheet.cell(row, COL_AVG_HOURS, avg_hours)
        sheet.cell(row, COL_PAY_BASE_DAYS, person["days"])
        sheet.cell(row, COL_TAXABLE, person["taxable"])
        sheet.cell(row, COL_WAGE_TOTAL, person["wage"])

        if insurance in ("3", "5"):
            sheet.cell(row, COL_RESIGN, resign)

        if report_nts:
            sheet.cell(row, COL_NTS_YN, "Y")
            sheet.cell(row, COL_PAY_MONTH, person["paymonth"] or site_paymonth or "")
            sheet.cell(row, COL_NTS_TOTAL, person["taxable"])
            if person["exempt"]:
                sheet.cell(row, COL_NTS_EXEMPT, person["exempt"])
            if person["tax"]:
                sheet.cell(row, COL_NTS_TAX, person["tax"])
            if person["local_tax"]:
                sheet.cell(row, COL_NTS_LOCAL_TAX, person["local_tax"])

        if foreign:
            foreigners.append(person["name"])

    workbook.save(out_path)

    return {
        "count": len(people),
        "days": sum(p["days"] for p in people),
        "wage": sum(p["wage"] for p in people),
        "foreigners": foreigners,
        "missing_days": missing_days,
        "day_mismatch": day_mismatch,
        "paymonth": site_paymonth,
        "path": out_path,
    }


def summarize_sites(sites):
    """생성 전 미리보기용 요약. 파일을 만들지 않는다."""
    summaries = []
    for site, people in sites.items():
        summaries.append({
            "site": site,
            "paymonth": next((p["paymonth"] for p in people if p["paymonth"]), None),
            "count": len(people),
            "days": sum(p["days"] for p in people),
            "wage": sum(p["wage"] for p in people),
            "foreigners": [p["name"] for p in people if is_foreigner(p["jumin"], p["nat_flag"])],
            "missing_days": [p["name"] for p in people if not p["day_marks"]],
            "day_mismatch": [
                p["name"] for p in people
                if p["day_marks"] and sum(1 for m in p["day_marks"] if m) != p["days"]
            ],
        })
    return summaries


def load_people(ledger_paths, grid_paths=()):
    """대장 여러 개를 읽어 (인원목록, 읽은내역, 합쳐진중복건수)를 돌려준다.

    같은 사람이 여러 파일에 나오면 한 사람으로 합치고, 일자별 출근표가 있는 값을 살린다.
    grid_paths 는 일자 칸 보완 전용(인원 수에는 더하지 않음)으로 쓴다.
    """
    people = []
    loaded = []
    for path in ledger_paths:
        fmt, parsed = parse_ledger(path)
        loaded.append({"path": path, "format": fmt, "count": len(parsed)})
        people.extend(parsed)

    people, duplicates = dedupe_people(people)

    grid_people = []
    for path in grid_paths:
        fmt, parsed = parse_ledger(path)
        if fmt != _FORMAT_A:
            loaded.append({"path": path, "format": fmt, "count": 0, "skipped": True})
            continue
        loaded.append({"path": path, "format": fmt, "count": len(parsed), "grid": True})
        grid_people.extend(parsed)

    if grid_people:
        fill_missing_day_marks(people, grid_people)

    return people, loaded, duplicates


def build_reports(ledger_paths, template_path, out_dir, params, grid_paths=()):
    """대장 → 현장별 신고서 파일 생성. (현장별결과, 읽은내역, 합쳐진중복건수)를 돌려준다."""
    people, loaded, duplicates = load_people(ledger_paths, grid_paths)
    if not people:
        raise LedgerFormatError("읽어들인 인원이 없습니다. 파일 내용을 확인해주세요.")

    os.makedirs(out_dir, exist_ok=True)

    results = []
    for site, site_people in group_by_site(people).items():
        paymonth = next((p["paymonth"] for p in site_people if p["paymonth"]), None)
        out_path = os.path.join(out_dir, report_filename(site, paymonth))
        result = write_site_report(template_path, out_path, site_people, params)
        result["site"] = site
        results.append(result)

    return results, loaded, duplicates
