"""
숙소보증금 현황표 생성 모듈
원장(임차보증금 계정별 원장) + 인사기록부 → 현황표 엑셀 자동 생성
"""

import os
import re
import sys
import threading
import traceback
from datetime import date, datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    import tkinter as _tk
    from tkinter import messagebox as _mb
    _r = _tk.Tk(); _r.withdraw()
    _mb.showerror("라이브러리 누락", f"필요한 라이브러리:\n{e}\n\npip install openpyxl pandas")
    sys.exit(1)

try:
    from tkinterdnd2 import DND_FILES
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False


# ──────────────────────────────────────────────
# 공통 유틸
# ──────────────────────────────────────────────

def _to_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except Exception:
        return 0


def _normalize(s):
    """공백·괄호 접미사 제거 후 비교용 문자열 반환.
    예) '장 성 호(전기)' → '장성호',  '0997.0' → '0997'
    """
    s = str(s or "").strip()
    s = re.sub(r"\(.*?\)", "", s)   # (전기), (57') 등 제거
    s = s.replace(" ", "").strip()
    return s


def _looks_like_name(s):
    """사람 이름처럼 보이는 문자열인지. '홍길동', "이상원(65년)" 앞부분 등은 True,
    '괴산군 괴산읍 동부리 522-5' 같은 주소는 False."""
    s = str(s or "").strip()
    if not s or len(s) > 6:
        return False
    if re.search(r"[0-9\s]", s):
        return False
    return bool(re.fullmatch(r"[가-힣]{2,6}", s))


# 프로젝트명 뒤에 상투적으로 붙는 어구 — 이 지점부터 잘라낸다
_SITE_CUT = re.compile(
    r"\s*(?:시공단계\s*)?(?:감독\s*권한대행|건설사업관리|통합건설사업관리|책임감리|감리용역|감리\s*용역)"
)
# 사업 코드 접미사 — (A233), (D082) 등은 현장 구분에 필요하므로 보존한다
_SITE_CODE = re.compile(r"\(([A-Z]\d{3})\)\s*$")
_SITE_TAIL = re.compile(r"\s*(?:등\s*\d+건|외\s*\d+건|등|PQ)\s*$")


def _short_site(name, maxlen=50):
    """원장 프로젝트명 → 현황표용 짧은 현장명.
    '154kV 정읍S/S 토건공사 감독권한대행 등 건설사업관리용역 PQ' → '154kV 정읍S/S 토건공사'
    '(25사단)26-A-00부대 건설사업관리용역(A032)'                 → '(25사단)26-A-00부대(A032)'
    """
    s = str(name or "").strip()
    if not s:
        return ""

    code = ""
    m = _SITE_CODE.search(s)
    if m:
        code = f"({m.group(1)})"
        s = s[:m.start()].strip()

    cut = _SITE_CUT.search(s)
    if cut and cut.start() > 0:
        s = s[:cut.start()].strip()

    # '등 6건', '외 1건', 'PQ' 같은 꼬리는 반복해서 제거
    prev = None
    while prev != s:
        prev = s
        s = _SITE_TAIL.sub("", s).strip()

    s = re.sub(r"\s+", " ", s).strip(" ,-/")
    s = (s + code) if code else s
    return s[:maxlen]


def _sabun_str(v):
    """사번을 4자리 문자열로 정규화. '1736.0' → '1736', '997' → '0997'"""
    if v in (None, "", " "):
        return ""
    s = str(v).strip().split(".")[0]
    if s.isdigit():
        return s.zfill(4)
    return s


# ──────────────────────────────────────────────
# 인사기록부 파싱
# ──────────────────────────────────────────────

def parse_hr(path, log_fn=print):
    """
    인사기록부(xls/xlsx) → (hr_dict, 퇴사자_set)
    hr_dict  : {사번: {이름, 소속, 직종, 재직여부}}
    퇴사자_set: 퇴사자 이름_norm 집합 (사번 없는 경우 대비)
    """
    if not path or not os.path.isfile(path):
        return {}, set()

    hr = {}
    퇴사자_names = set()

    def _read_sheet(ws_rows, is_퇴사자=False):
        hdr_row = 0
        for r_idx, row in enumerate(ws_rows):
            vals = [str(c or "").replace(" ", "") for c in row]
            if "사번" in vals and ("성명" in vals or "이름" in vals):
                hdr_row = r_idx
                break
        if not ws_rows:
            return

        col_sabun = col_name = col_dept = col_jikjong = None
        for c_idx, v in enumerate(ws_rows[hdr_row]):
            v_s = str(v or "").replace(" ", "")
            if v_s == "사번":                           col_sabun   = c_idx
            elif v_s in ("성명", "이름"):               col_name    = c_idx
            elif "소" in v_s and "속" in v_s:           col_dept    = c_idx
            elif v_s == "직종":                         col_jikjong = c_idx

        if col_name is None:
            return

        for row in ws_rows[hdr_row + 1:]:
            try:
                sabun = _sabun_str(row[col_sabun]) if col_sabun is not None else ""
                name  = _normalize(str(row[col_name] or ""))
                dept  = str(row[col_dept] or "").strip() if col_dept is not None else ""
                jj    = str(row[col_jikjong] or "").strip() if col_jikjong is not None else ""
                if not name:
                    continue
                if is_퇴사자:
                    퇴사자_names.add(name)
                    if sabun:
                        hr[sabun] = {"이름": name, "소속": dept, "직종": jj, "퇴사": True}
                else:
                    if sabun:
                        hr[sabun] = {"이름": name, "소속": dept, "직종": jj, "퇴사": False}
            except Exception:
                pass

    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(path)
            for sname in ["(주)선종합", "Sheet1", "재직자", wb.sheet_names()[0]]:
                if sname in wb.sheet_names():
                    ws = wb.sheet_by_name(sname)
                    rows = [ws.row_values(r) for r in range(ws.nrows)]
                    _read_sheet(rows, is_퇴사자=False)
                    break
            if "퇴사자명부" in wb.sheet_names():
                ws2 = wb.sheet_by_name("퇴사자명부")
                rows2 = [ws2.row_values(r) for r in range(ws2.nrows)]
                _read_sheet(rows2, is_퇴사자=True)
        else:
            owb = load_workbook(path, data_only=True)
            sheet_names = owb.sheetnames
            for sname in sheet_names:
                is_q = "퇴사" in sname
                ws = owb[sname]
                rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                        for r in range(1, ws.max_row + 1)]
                _read_sheet(rows, is_퇴사자=is_q)
    except Exception as e:
        log_fn(f"   ⚠ 인사기록부 로딩 오류: {e}")
        return {}, set()

    log_fn(f"   인사기록부: 재직 {sum(1 for v in hr.values() if not v['퇴사'])}명 / 퇴사 {len(퇴사자_names)}명")
    return hr, 퇴사자_names


# ──────────────────────────────────────────────
# 원장 파싱
# ──────────────────────────────────────────────

def _find_header_row(ws):
    """헤더 행 번호(1-indexed) 반환 — 날짜/차변 컬럼으로 감지."""
    for r in range(1, 15):
        vals = [str(ws.cell(r, c).value or "") for c in range(1, 8)]
        joined = "".join(vals)
        if "날짜" in joined or ("차변" in joined and "대변" in joined):
            return r
    return 1


def _date_str_from(val, year=None):
    """openpyxl 날짜 값(datetime / 문자열 / 숫자) → 'YYYY-MM-DD' 문자열."""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # 이미 YYYY-MM-DD 형태면 그대로
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # YYYYMMDD 8자리 숫자
    if re.match(r"^\d{8}$", s):
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    # MM-DD (연도 없음 → 파일명 연도 사용)
    m = re.match(r"^(\d{1,2})-(\d{1,2})$", s)
    if m:
        y = year or date.today().year
        return f"{y}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return s


def _year_from_filename(path):
    """파일명에서 연도 추출. 예) '원장(2025)' → 2025, '원장(2026.07)' → 2026."""
    name = os.path.basename(path)
    m = re.search(r'\((\d{4})', name)
    if m:
        return int(m.group(1))
    m2 = re.search(r'(\d{4})', name)
    if m2:
        yr = int(m2.group(1))
        if 2000 <= yr <= 2100:
            return yr
    return date.today().year


def _ledger_closing_balance(ws, hdr, col):
    """원장의 마지막 데이터 행 잔액 → int. 못 찾으면 None.
    [월 계]/[누 계] 행은 잔액이 비어 있으므로 자연히 건너뛴다."""
    c_bal, c_date = col.get("잔액"), col.get("날짜")
    if not c_bal or not c_date:
        return None
    for r in range(ws.max_row, hdr, -1):
        if not ws.cell(r, c_date).value:
            continue
        v = ws.cell(r, c_bal).value
        if v not in (None, ""):
            return _to_int(v)
    return None


def parse_ledger(path, log_fn=print):
    """
    원장 파일 → (거래 내역 리스트, 마지막 잔액)
    각 항목: {date_str, 적요, 사번, 이름, 차변, 대변, 거래처, 프로젝트명, 거래유형}
    """
    ledger_year = _year_from_filename(path)
    log_fn(f"   연도 감지: {ledger_year}년  ({os.path.basename(path)})")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    hdr = _find_header_row(ws)
    log_fn(f"   헤더 행: {hdr}행")

    # 헤더명으로 컬럼 위치 탐지.
    # 사번·이름·프로젝트명은 위치를 고정하지 않는다 — 사원코드 열이 없는 원장에서
    # 프로젝트명 열을 사번으로 잘못 읽는 문제가 있었다.
    col = {}
    for c in range(1, ws.max_column + 1):
        v  = _normalize(ws.cell(hdr, c).value)
        vl = v.lower()
        if   "일자" in v or "날짜" in v:              col.setdefault("날짜", c)
        elif "적요" in v:                             col.setdefault("적요", c)
        elif "거래처" in v:                           col.setdefault("거래처", c)
        elif "차변" in v:                             col.setdefault("차변", c)
        elif "대변" in v:                             col.setdefault("대변", c)
        elif "잔액" in v:                             col.setdefault("잔액", c)
        elif vl == "code" or v in ("사번", "사원코드"): col.setdefault("사번", c)
        elif v in ("사원명", "성명", "이름"):          col.setdefault("이름", c)
        elif "프로젝트" in v:                         col.setdefault("프로젝트명", c)

    col.setdefault("날짜", 1)
    col.setdefault("적요", 2)
    col.setdefault("차변", 4)
    col.setdefault("대변", 5)
    log_fn("   컬럼: " + ", ".join(
        f"{k}={get_column_letter(v)}" for k, v in sorted(col.items(), key=lambda x: x[1])))
    for k in ("사번", "이름", "프로젝트명"):
        if k not in col:
            log_fn(f"   ⚠ '{k}' 열 없음 — 적요에서 추출하거나 공란 처리")

    def _cell(r, key):
        c = col.get(key)
        return ws.cell(r, c).value if c else None

    # 날짜 샘플 출력 (첫 3건)
    _sample = 0
    txns = []
    for r in range(hdr + 1, ws.max_row + 1):
        date_val = _cell(r, "날짜")
        적요_val  = _cell(r, "적요")
        차변_val  = _cell(r, "차변")
        대변_val  = _cell(r, "대변")
        사번_val  = _cell(r, "사번")
        이름_val  = _cell(r, "이름")
        거래처_val = _cell(r, "거래처")
        proj_val  = _cell(r, "프로젝트명")

        if not date_val:
            continue
        적요_str = str(적요_val or "")
        if any(k in 적요_str for k in ["[", "이월", "소 계", "누 계", "소계", "누계"]):
            continue

        차변 = _to_int(차변_val)
        대변 = _to_int(대변_val)
        if 차변 == 0 and 대변 == 0:
            continue

        if "이전" in 적요_str:
            유형 = "이전"
        elif 차변 > 0 and 대변 == 0:
            유형 = "지급"
        elif 대변 > 0 and 차변 == 0:
            유형 = "반환"
        elif 차변 > 0 and 대변 > 0:
            유형 = "이전"
        else:
            유형 = "기타"

        if 유형 == "기타":
            continue

        거래처_str = str(거래처_val or "").strip()
        # 거래처가 '직원'이 아니면 임대인·법인 명의 계약(현장사무실 등)
        is_사무실 = bool(거래처_str) and 거래처_str != "직원"

        사번_str = _sabun_str(사번_val)
        이름_str = str(이름_val or "").strip()
        if not 이름_str and not is_사무실:
            # 적요의 '보증금(홍길동)' 패턴에서만 이름 추출.
            # 임대인 계약은 괄호 안이 주소·임대인이라 사람 이름으로 오인하면 안 된다.
            m = re.search(r"(?:지급|반환|이전|보증금)\s*\(([^)]{2,12})\)", 적요_str)
            if m and _looks_like_name(m.group(1)):
                이름_str = m.group(1).strip()

        date_str = _date_str_from(date_val, ledger_year)
        if _sample < 3:
            log_fn(f"   [날짜샘플] raw={repr(date_val)} → '{date_str}'")
            _sample += 1

        txns.append({
            "date_str":   date_str,
            "적요":       적요_str,
            "사번":       사번_str,
            "이름":       이름_str,
            "이름_norm":  _normalize(이름_str),
            "차변":       차변,
            "대변":       대변,
            "거래처":     거래처_str,
            "사무실":     is_사무실,
            "프로젝트명": str(proj_val or "").strip(),
            "거래유형":   유형,
        })

    balance = _ledger_closing_balance(ws, hdr, col)
    log_fn(f"   유효 거래: {len(txns)}건"
           + (f" / 최종 잔액 {balance:,}원" if balance is not None else ""))
    return txns, balance


# ──────────────────────────────────────────────
# 현황표 로딩
# ──────────────────────────────────────────────

_COL_ALIAS = {
    "번호":        ["번호", "구분", "no", "No"],
    "현장명":      ["현장명", "현 장 명"],
    "보증금지급액": ["보증금지급액", "보증금", "지급액", "임차보증금"],
    "지급일자":    ["지급일자", "지 급 일 자", "계약기간", "계 약 기 간"],
    "사번":        ["사번", "code", "사원코드"],
    "지급자":      ["지급자", "지급(보증)자", "지급보증자", "상주감리원"],
    "현황":        ["현황", "비고"],
    "회계처리":    ["회계처리", "회수일"],
}

def _find_col(df_cols, aliases):
    for alias in aliases:
        for c in df_cols:
            if _normalize(c) == _normalize(alias):
                return c
    return None


def load_status(path, log_fn=print):
    """기존 현황표 → DataFrame (내부 컬럼명 통일)"""
    df = pd.read_excel(path)
    log_fn(f"   원본 컬럼: {list(df.columns)}")

    # 이중 헤더 처리: 첫 번째 데이터 행이 서브헤더인 경우 제거
    # (예: 컬럼명='계 약 기 간', 첫 행 값='지급일')
    if len(df) > 0:
        first_row = df.iloc[0]
        non_null = first_row.dropna()
        # 첫 행의 값이 모두 문자열이고 숫자가 없으면 서브헤더로 판단
        if len(non_null) > 0 and all(
            isinstance(v, str) and not re.match(r'^\d{5,}$', str(v))
            for v in non_null.values
        ):
            log_fn(f"   이중 헤더 감지 → 첫 행 제거: {list(non_null.values)}")
            df = df.iloc[1:].reset_index(drop=True)

    rename = {}
    for key, aliases in _COL_ALIAS.items():
        c = _find_col(df.columns.tolist(), aliases)
        if c and c != key:
            rename[c] = key
    if rename:
        log_fn(f"   컬럼 매핑: {rename}")
    else:
        log_fn(f"   컬럼 매핑: (변경 없음 — 원본 컬럼명과 동일)")
    df = df.rename(columns=rename)

    if "현장명" in df.columns:
        df = df[df["현장명"].notna()].copy()
        # '소     계'처럼 글자 사이에 공백이 여러 개 들어간 합계행도 걸러야 한다.
        # 공백을 모두 지운 뒤 비교하지 않으면 합계 금액이 데이터 한 줄로 살아남는다.
        _site_flat = df["현장명"].astype(str).str.replace(r"\s+", "", regex=True)
        _is_total  = _site_flat.str.contains("합계|소계|누계|총계", na=False)
        if _is_total.any():
            log_fn(f"   합계행 제외: {_is_total.sum()}행")
        df = df[~_is_total].copy()

    # 보증금지급액이 0이거나 비어있는 행 경고
    if "보증금지급액" in df.columns:
        empty_amt = df["보증금지급액"].apply(_to_int) == 0
        if empty_amt.any():
            log_fn(f"   ⚠ 보증금지급액 없는 행 {empty_amt.sum()}건 (원장에서 매칭 시 제외)")

    if "사번" not in df.columns:
        df["사번"] = ""

    for col in ["번호", "현장명", "보증금지급액", "지급일자", "사번", "지급자", "현황", "회계처리"]:
        if col not in df.columns:
            df[col] = ""

    if "지급자" in df.columns:
        df["지급자_norm"] = df["지급자"].apply(lambda x: _normalize(str(x or "")))
    else:
        df["지급자_norm"] = ""

    df["사번"] = df["사번"].apply(lambda x: _sabun_str(x) if x else "")

    df = df.reset_index(drop=True)
    log_fn(f"   기존 현황표: {len(df)}행 로딩")
    return df


# ──────────────────────────────────────────────
# 업데이트 로직
# ──────────────────────────────────────────────

def _date_to_sortkey(val):
    """날짜 값 → 정렬/비교용 YYYYMMDD 정수. 0이면 파싱 실패."""
    if val is None or val == "" or (hasattr(val, '__class__') and val.__class__.__name__ == 'NaTType'):
        return 0

    # datetime / pandas Timestamp / date 객체
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
        try:
            return val.year * 10000 + val.month * 100 + val.day
        except Exception:
            pass

    s = str(val).strip()

    # "NaT", "nan", "None"
    if s.lower() in ("nat", "nan", "none", ""):
        return 0

    # YYYY-MM-DD  또는  YYYY-MM-DD HH:MM:SS
    m = re.match(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        return int(m.group(1)) * 10000 + int(m.group(2)) * 100 + int(m.group(3))

    # YYYYMMDD (8자리 숫자)
    m2 = re.match(r"^(\d{8})$", s)
    if m2:
        return int(s)

    # MM-DD 또는 MM/DD (연도 없음 → 그대로 0으로 두고 비교 불가 처리)
    m3 = re.match(r"^(\d{1,2})[-/](\d{1,2})$", s)
    if m3:
        # 연도 정보 없으므로 비교 불가 → 0 반환
        return 0

    # 순수 숫자(Excel 날짜 시리얼)
    try:
        n = float(s)
        if 30000 < n < 60000:
            # Excel 1900 epoch
            from datetime import timedelta
            epoch = datetime(1899, 12, 30)
            d = epoch + timedelta(days=n)
            return d.year * 10000 + d.month * 100 + d.day
        return int(n)
    except Exception:
        return 0


def _site_key(name):
    """현장명 비교용 키 — 공백·기호 제거"""
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(name or "")).lower()


def _same_site(a, b, minlen=4):
    """두 현장명이 같은 현장을 가리키는지 — 앞부분 공통 접두 길이로 판단.
    '센트럴밸리산업단지현장사무실(312호)' 와
    '센트럴밸리 산업단지 용수공급시설 설치사업' 을 같은 현장으로 본다."""
    ka, kb = _site_key(a), _site_key(b)
    if not ka or not kb:
        return False
    if ka == kb:
        return True
    n = 0
    for x, y in zip(ka, kb):
        if x != y:
            break
        n += 1
    return n >= minlen


_사무실_지급자 = ("", "법인")


def _find_target_row(rows, txn, 금액, log_fn=None):
    """현황표에서 반환·이전 대상 행 찾기. rows 는 dict 리스트(_alive 플래그 사용).
    반환값: (행 인덱스, 금액일치여부). 못 찾으면 (None, False)."""
    사번      = txn["사번"]
    이름_norm = txn["이름_norm"]
    live = [i for i, r in enumerate(rows) if r["_alive"]]

    def amt(i):
        return _to_int(rows[i]["보증금지급액"])

    # 1순위: 사번 + 금액
    if 사번:
        for i in live:
            if rows[i]["사번"] == 사번 and amt(i) == 금액:
                return i, True
    # 2순위: 이름 + 금액
    if 이름_norm:
        for i in live:
            if _name_hit(rows[i]["지급자_norm"], txn) and amt(i) == 금액:
                return i, True

    # 3순위: 임대인·법인 명의(사원명 없음) — 현장명 + 금액으로 찾는다
    if not 이름_norm:
        for i in live:
            if (rows[i]["지급자_norm"] in _사무실_지급자 and amt(i) == 금액
                    and (_same_site(rows[i]["현장명"], txn["프로젝트명"])
                         or _same_site(rows[i]["현장명"], txn["적요"]))):
                return i, True
        # 4순위: 금액만 일치하는 임대인 행 (FIFO)
        for i in live:
            if rows[i]["지급자_norm"] in _사무실_지급자 and amt(i) == 금액:
                return i, True
        return None, False

    # 5순위: 이름만 일치 — 금액이 어긋난다.
    #        원장에는 한 현장의 보증금 일부만 반환·이전하는 건이 있으므로
    #        (예: 10,000,000원 중 5,000,000원만 다른 현장으로 이전)
    #        반환액보다 큰 행을 우선 골라 부분 반환으로 처리할 수 있게 한다.
    for i in live:
        if _name_hit(rows[i]["지급자_norm"], txn) and amt(i) > 금액:
            return i, False
    for i in live:
        if _name_hit(rows[i]["지급자_norm"], txn):
            return i, False
    return None, False


def _pair_transfers(txns, log_fn=print):
    """'이전' 거래의 차변·대변 짝을 맞춘다.

    같은 날짜·같은 사람·같은 금액의 차변 1건 + 대변 1건 = 현장 이동 1건.
    짝이 없는 이전은 실제로는 신규 지급이거나 반환이므로 각각으로 되돌린다.
    (예: '보증금이전및지급(박찬수) 10,000,000' 은 짝이 없는 지급이다)
    """
    # 원장에 적힌 순서를 보존한다. 같은 날짜에 '이전 → 반환'이 이어지는 건이 있어
    # 순서가 뒤집히면 잔액 추적이 어긋난다.
    for seq, t in enumerate(txns):
        t.setdefault("_seq", seq)

    out, 이전 = [], []
    for t in txns:
        (이전 if t["거래유형"] == "이전" else out).append(t)

    used = set()
    for i, a in enumerate(이전):
        if i in used or a["차변"] <= 0 or a["대변"] > 0:
            continue
        key_a = (a["date_str"], a["사번"] or a["이름_norm"], a["차변"])
        for j, b in enumerate(이전):
            if j in used or j == i or b["대변"] <= 0 or b["차변"] > 0:
                continue
            if key_a == (b["date_str"], b["사번"] or b["이름_norm"], b["대변"]):
                used.add(i)
                used.add(j)
                t = dict(a)
                t["이전전_프로젝트명"] = b["프로젝트명"]
                t["_짝맞춘이전"] = True     # 잔액 변동 없음 (현장만 이동)
                out.append(t)
                break

    for i, t in enumerate(이전):
        if i in used:
            continue
        t = dict(t)
        if t["차변"] > 0 and t["대변"] == 0:
            t["거래유형"] = "지급"
        elif t["대변"] > 0 and t["차변"] == 0:
            t["거래유형"] = "반환"
        else:
            out.append(t)
            continue
        who = t["이름"] or t["거래처"]
        log_fn(f"   ℹ 이전 짝 없음 → {t['거래유형']}으로 처리: "
               f"{who} {max(t['차변'], t['대변']):,}원")
        out.append(t)

    out.sort(key=lambda t: (_date_to_sortkey(t["date_str"]), t.get("_seq", 0)))
    return out


def _new_row(txn, 금액, 이름_raw, 이름_norm):
    """원장 거래 1건 → 현황표 신규 행"""
    is_사무실 = bool(txn.get("사무실"))
    현장명 = _short_site(txn["프로젝트명"]) or _short_site(txn["적요"])
    return {
        "번호":        None,
        "현장명":      현장명,
        "보증금지급액": 금액,
        "지급일자":    txn["date_str"],
        "사번":        txn["사번"],
        "지급자":      "" if is_사무실 else 이름_raw,
        "지급자_norm": "" if is_사무실 else 이름_norm,
        # 비고(대여금 등)는 회계 판단이라 자동으로 채우지 않는다
        "현황":        "",
        "회계처리":    "",
        "_alive":      True,
    }


def find_missing_rows(rows, past_txns, log_fn=print):
    """기준일 이전 원장에서 '지급됐고 아직 반환되지 않은' 건을 뽑아,
    현황표에 없는 것만 골라 반환한다.

    기준파일이 수기로 관리돼 원장에 있는 건이 누락돼 있을 때 이를 찾아낸다.
    기준파일의 지급일자는 일괄 기입된 값이 많아 날짜로는 대조할 수 없으므로
    사번·이름·금액으로만 판단한다.
    """
    # 원장 기준 사람별 미반환 잔액
    open_pos = []
    for t in past_txns:
        if t.get("_짝맞춘이전"):
            continue                     # 현장만 옮긴 건 — 잔액은 그대로다
        금액 = t["차변"] or t["대변"]
        키   = t["이름_norm"] or f"[{t['거래처']}]"
        if t["차변"] > 0:
            open_pos.append((키, 금액, t))
        else:
            for p in open_pos:
                if p[0] == 키 and p[1] == 금액:
                    open_pos.remove(p)
                    break
            else:
                # 금액이 딱 맞는 건이 없으면 같은 사람의 가장 오래된 건에서 차감
                for p in open_pos:
                    if p[0] == 키:
                        open_pos.remove(p)
                        open_pos.append((키, p[1] - 금액, p[2]))
                        break

    ledger_total = {}
    latest_txn   = {}
    for 키, 금액, t in open_pos:
        ledger_total[키] = ledger_total.get(키, 0) + 금액
        latest_txn[키] = t

    # 기준파일은 한 사람의 여러 지급건을 한 행으로 합쳐 둔 경우가 많아
    # 건별로는 대조할 수 없다. 사람별 총액이 원장보다 모자란 만큼만 보충한다.
    # (원장에 없는 과거 지급분이 있을 수 있으므로 반대 방향은 건드리지 않는다)
    missing, 이름의심 = [], []
    for 키, led in sorted(ledger_total.items()):
        t  = latest_txn[키]
        후보 = t.get("이름_후보") or []
        sheet = sum(_to_int(r["보증금지급액"]) for r in rows
                    if r["_alive"] and r["지급자_norm"] in 후보) if 후보 else 0
        부족 = led - sheet
        if 부족 <= 0:
            continue

        # 임대인 명의는 현장명으로도 한 번 더 확인 (지급자명이 비어 있을 수 있다)
        if not t["이름_norm"]:
            covered = sum(_to_int(r["보증금지급액"]) for r in rows
                          if r["_alive"] and r["지급자_norm"] in _사무실_지급자
                          and _same_site(r["현장명"], t["프로젝트명"]))
            부족 = led - covered
            if 부족 <= 0:
                continue

        # 금액·지급일자가 똑같은 행이 이미 있으면 이름 표기만 다른 같은 건으로 본다.
        # (기준파일에 '김봉희', 원장에 '김봉회'처럼 한 글자 다른 사례가 있다)
        dup = next((r for r in rows
                    if r["_alive"]
                    and _to_int(r["보증금지급액"]) == 부족
                    and _date_to_sortkey(r["지급일자"]) == _date_to_sortkey(t["date_str"])
                    and _date_to_sortkey(t["date_str"]) > 0), None)
        if dup is not None:
            이름의심.append({
                "일자": t["date_str"], "구분": "이름 표기 불일치 의심",
                "지급자": t["이름"] or t["거래처"], "금액": 부족, "적요": t["적요"],
                "프로젝트명": f"현황표에는 '{dup['지급자']}' / {dup['현장명']} 으로 기재 — 같은 건인지 확인",
            })
            log_fn(f"   ⚠ 이름 표기 불일치 의심: 원장 '{t['이름']}' vs 현황표 "
                   f"'{dup['지급자']}'  {부족:,}원 {t['date_str']} → 보충하지 않음")
            continue

        missing.append((t, 부족))
    return missing, 이름의심


def _apply_hr_names(txns, hr, log_fn=print):
    """인사기록부 이름으로 원장 이름을 보정한다 (사번이 있을 때만).

    원장 이름을 덮어쓰지 않고 '이름_후보'에 둘 다 담는다. 기존 현황표가 원장·인사
    기록부와 다른 표기를 쓰는 경우가 있어(류상완/유상완), 어느 쪽으로도 매칭돼야 한다.
    """
    seen = set()
    for t in txns:
        후보 = [t["이름_norm"]] if t["이름_norm"] else []
        사번 = t["사번"]
        if hr and 사번 and 사번 in hr:
            hr_name = hr[사번]["이름"]
            if hr_name and hr_name not in 후보:
                if 후보 and 사번 not in seen:
                    log_fn(f"   ℹ 이름 표기 차이: 원장 '{t['이름']}' / 인사기록부 "
                           f"'{hr_name}'  (사번 {사번}) — 둘 다로 대조")
                    seen.add(사번)
                후보.append(hr_name)
                if not t["이름_norm"]:          # 원장에 이름이 없으면 인사기록부로 채움
                    t["이름"] = hr_name
                    t["이름_norm"] = hr_name
        t["이름_후보"] = 후보
    return txns


def _name_hit(row_name_norm, txn):
    """현황표 행의 지급자가 이 거래의 사람과 같은지"""
    if not row_name_norm:
        return False
    return row_name_norm in txn.get("이름_후보", [txn["이름_norm"]])


def update_status(df, txns, hr=None, log_fn=print, past_txns=None):
    """기존 현황표 + 원장 → (새 현황표 DataFrame, 처리 리포트).

    hr        : parse_hr() 반환값의 첫 번째 요소 (사번→{이름,...}), None이면 미사용.
    past_txns : 기준일 이전 거래. 주면 현황표에 빠진 미반환 건을 찾아 채운다.

    행을 DataFrame 대신 dict 리스트로 다룬다. 신규 추가 행을 곧바로 목록에 넣어야
    같은 원장 안에서 지급 후 반환된 건("2월 지급 → 5월 반환")도 매칭된다.
    """
    added = removed = transferred = 0
    unmatched = []
    backfilled = []

    rows = df.to_dict("records")
    for r in rows:
        r["_alive"] = True
        if not r.get("지급자_norm"):
            r["지급자_norm"] = _normalize(str(r.get("지급자") or ""))

    txns = _apply_hr_names(_pair_transfers(txns, log_fn), hr, log_fn)

    # ── 기준일 이전 누락 건 채우기 ──
    # 기준일 이후 거래를 적용하기 전에 해야, 이후의 반환 거래가 제대로 매칭된다.
    if past_txns:
        past = _apply_hr_names(_pair_transfers(list(past_txns), lambda *a: None),
                               hr, lambda *a: None)
        missing, 이름의심 = find_missing_rows(rows, past, log_fn)
        unmatched.extend(이름의심)
        for t, 금액 in missing:
            row = _new_row(t, 금액, t["이름"], t["이름_norm"])
            row["현황"] = "원장대조 추가"
            rows.append(row)
            backfilled.append({
                "일자": t["date_str"], "구분": "누락 보충",
                "지급자": t["이름"] or t["거래처"], "금액": 금액,
                "적요": t["적요"], "프로젝트명": t["프로젝트명"],
            })
            log_fn(f"   ⊕ 누락 보충: {t['이름'] or t['거래처']}  {금액:,}원  {row['현장명'][:24]}")
        if backfilled:
            log_fn(f"   기준일 이전 누락 {len(backfilled)}건 보충 "
                   f"({sum(b['금액'] for b in backfilled):,}원)")

    for txn in txns:
        유형      = txn["거래유형"]
        사번      = txn["사번"]
        이름_raw  = txn["이름"]
        이름_norm = txn["이름_norm"]
        차변      = txn["차변"]
        대변      = txn["대변"]
        proj      = txn["프로젝트명"]
        date_str  = txn["date_str"]

        표기명 = 이름_raw or txn.get("거래처") or ""

        if 유형 == "지급":
            금액 = 차변
            date_key = _date_to_sortkey(date_str)
            already = False
            for r in rows:
                if not r["_alive"]:
                    continue
                name_ok = (사번 and r["사번"] == 사번) or _name_hit(r["지급자_norm"], txn)
                if (name_ok
                        and _to_int(r["보증금지급액"]) == 금액
                        and _date_to_sortkey(r["지급일자"]) == date_key):
                    already = True
                    if 사번 and not r["사번"]:
                        r["사번"] = 사번
                    break

            if not already:
                row = _new_row(txn, 금액, 이름_raw, 이름_norm)
                rows.append(row)
                added += 1
                log_fn(f"   + 추가: {표기명}  {금액:,}원  {row['현장명'][:28]}")

        elif 유형 == "반환":
            금액 = 대변
            idx, amt_ok = _find_target_row(rows, txn, 금액)
            if idx is not None:
                row_amt = _to_int(rows[idx]["보증금지급액"])
                현장 = rows[idx]["현장명"][:24]
                if amt_ok or row_amt <= 금액:
                    rows[idx]["_alive"] = False
                    removed += 1
                    log_fn(f"   - 반환삭제: {표기명}  {금액:,}원  {현장}")
                else:
                    # 부분 반환 — 행을 지우지 않고 남은 금액만 줄인다
                    rows[idx]["보증금지급액"] = row_amt - 금액
                    removed += 1
                    log_fn(f"   - 부분반환: {표기명}  {금액:,}원  {현장}  "
                           f"({row_amt:,} → {row_amt - 금액:,}원)")
                if not amt_ok:
                    unmatched.append({
                        "일자": date_str,
                        "구분": "부분반환" if row_amt > 금액 else "반환(금액불일치)",
                        "지급자": 표기명, "금액": 금액, "적요": txn["적요"],
                        "프로젝트명": f"현황표 {rows[idx]['현장명']} / 기존 {row_amt:,}원 → 확인 필요",
                    })
            else:
                unmatched.append({
                    "일자": date_str, "구분": "반환", "지급자": 표기명,
                    "금액": 금액, "적요": txn["적요"], "프로젝트명": proj,
                })
                log_fn(f"   ⚠ 반환 매칭 실패: {표기명}  {금액:,}원  → 수동 확인 필요")

        elif 유형 == "이전":
            금액 = 차변
            idx, _amt_ok = _find_target_row(rows, txn, 금액)
            if idx is not None:
                old = rows[idx]["현장명"]
                rows[idx]["현장명"] = _short_site(proj) or old
                transferred += 1
                log_fn(f"   ↔ 이전: {표기명}  {old[:22]} → {rows[idx]['현장명'][:22]}")
            else:
                # 이동 대상이 표에 없으면 새로 지급된 것으로 본다
                row = _new_row(txn, 금액, 이름_raw, 이름_norm)
                rows.append(row)
                added += 1
                log_fn(f"   + 추가(이전 대상 없음): {표기명}  {금액:,}원  {row['현장명'][:24]}")

    live = [r for r in rows if r["_alive"]]
    for r in live:
        r.pop("_alive", None)
    df = pd.DataFrame(live, columns=_HEADERS + ["지급자_norm"])

    # 지급일자 정렬
    df["_sort"] = df["지급일자"].apply(_date_to_sortkey)
    df = df.sort_values("_sort", na_position="last").drop(columns=["_sort"])
    df = df.reset_index(drop=True)
    df["번호"] = range(1, len(df) + 1)

    total = df["보증금지급액"].apply(_to_int).sum()
    log_fn(f"\n   ── 처리 요약 ──")
    log_fn(f"   신규 추가:  {added}건")
    log_fn(f"   반환 삭제:  {removed}건")
    log_fn(f"   현장 이전:  {transferred}건")
    log_fn(f"   누락 보충:  {len(backfilled)}건")
    log_fn(f"   확인 필요:  {len(unmatched)}건")
    log_fn(f"   최종 행수:  {len(df)}건")
    log_fn(f"   총 보증금:  {total:,}원")

    report = {
        "added": added, "removed": removed, "transferred": transferred,
        "unmatched": unmatched, "backfilled": backfilled, "total": int(total),
    }
    return df, report


# ──────────────────────────────────────────────
# 저장
# ──────────────────────────────────────────────

_HEADERS  = ["번호", "현장명", "보증금지급액", "지급일자", "사번", "지급자", "현황", "회계처리"]
_COL_W    = [7,      42,       14,             13,         7,     10,     10,   12   ]

# 퇴사자 행 강조색
_FILL_퇴사  = PatternFill("solid", fgColor="FFD966")   # 노란색
_FILL_ALT   = PatternFill("solid", fgColor="F5F8FF")   # 연파랑 교번
_FILL_HDR   = PatternFill("solid", fgColor="1F497D")   # 진파랑 헤더
_FILL_TOTAL = PatternFill("solid", fgColor="E8EEF7")   # 합계행


def _normalize_date(val):
    """지급일자 값을 YYYY-MM-DD 문자열로 통일."""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    s = str(val or "").strip()
    # YYYY-MM-DD
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    # YYYYMMDD
    m2 = re.match(r"^(\d{4})(\d{2})(\d{2})$", s)
    if m2:
        return f"{m2.group(1)}-{m2.group(2)}-{m2.group(3)}"
    # YYYY.MM.DD
    m3 = re.match(r"(\d{4})\.(\d{2})\.(\d{2})", s)
    if m3:
        return f"{m3.group(1)}-{m3.group(2)}-{m3.group(3)}"
    return s


def _write_sheet(ws, df, 퇴사자_names, title_label):
    """df 를 ws 에 표 형식으로 쓴다."""
    thin   = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font    = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
    normal_font = Font(name="맑은 고딕", size=9)
    total_font  = Font(bold=True, name="맑은 고딕", size=9)
    퇴사_font   = Font(name="맑은 고딕", size=9, color="C00000")

    for c, (h, w) in enumerate(zip(_HEADERS, _COL_W), 1):
        cell = ws.cell(1, c, h)
        cell.fill = _FILL_HDR
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        이름_norm = _normalize(str(getattr(row, "지급자", "") or ""))
        사번      = str(getattr(row, "사번", "") or "")
        is_퇴사   = 이름_norm and 이름_norm in 퇴사자_names

        현황_val = "퇴사" if is_퇴사 else (str(getattr(row, "현황", "") or ""))
        date_val = _normalize_date(getattr(row, "지급일자", ""))

        vals = [
            getattr(row, "번호", r_idx - 1),
            getattr(row, "현장명", ""),
            _to_int(getattr(row, "보증금지급액", 0)) or None,
            date_val,
            사번,
            getattr(row, "지급자", ""),
            현황_val,
            getattr(row, "회계처리", ""),
        ]

        base_fill = _FILL_퇴사 if is_퇴사 else (_FILL_ALT if r_idx % 2 == 0 else None)

        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_idx, c, v)
            cell.font = 퇴사_font if is_퇴사 else normal_font
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c in (1, 4, 5, 7) else "left",
            )
            if base_fill:
                cell.fill = base_fill
            if c == 3 and v:
                cell.number_format = "#,##0"
        ws.row_dimensions[r_idx].height = 18

    # 합계 행
    total_row = len(df) + 2
    total_amt = sum(_to_int(getattr(r, "보증금지급액", 0)) for r in df.itertuples())
    ws.cell(total_row, 1, "합  계")
    ws.cell(total_row, 3, total_amt).number_format = "#,##0"
    for c in range(1, len(_HEADERS) + 1):
        cell = ws.cell(total_row, c)
        if c == 3:
            cell.value = total_amt
            cell.number_format = "#,##0"
        cell.font  = total_font
        cell.fill  = _FILL_TOTAL
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[total_row].height = 20
    ws.freeze_panes = "A2"
    return total_amt


def _write_check_sheet(ws, report, ledger_balance, total_amt, log_fn=print):
    """원장 잔액 대조 + 매칭 실패 목록 시트"""
    hdr_font   = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
    bold_font  = Font(bold=True, name="맑은 고딕", size=10)
    norm_font  = Font(name="맑은 고딕", size=9)
    warn_font  = Font(name="맑은 고딕", size=9, color="C00000")

    for c, w in enumerate([14, 10, 12, 14, 46, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.cell(1, 1, "원장 잔액 대조").font = bold_font
    diff = None
    lines = [("현황표 합계", total_amt)]
    if ledger_balance is not None:
        diff = total_amt - ledger_balance
        lines.append(("원장 최종 잔액", ledger_balance))
        lines.append(("차액 (현황표 − 원장)", diff))

    r = 2
    for label, val in lines:
        ws.cell(r, 1, label).font = norm_font
        cell = ws.cell(r, 3, val)
        cell.number_format = "#,##0"
        cell.font = warn_font if (label.startswith("차액") and val) else norm_font
        r += 1

    if diff:
        ws.cell(r, 1, "※ 차액은 원장에 없는 과거 잔액이거나 미매칭 건입니다. 아래 목록을 확인하세요.").font = warn_font
        r += 1
    r += 1

    cols = ["일자", "구분", "지급자", "금액", "적요", "프로젝트명"]

    def _table(title, items):
        nonlocal r
        ws.cell(r, 1, f"{title} ({len(items)}건)").font = bold_font
        r += 1
        for c, h in enumerate(cols, 1):
            cell = ws.cell(r, c, h)
            cell.fill = _FILL_HDR
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
        for item in items:
            for c, key in enumerate(cols, 1):
                cell = ws.cell(r, c, item.get(key, ""))
                cell.font = norm_font
                if key == "금액":
                    cell.number_format = "#,##0"
            r += 1
        r += 1

    _table("확인 필요 (매칭 실패·금액 불일치)", report.get("unmatched", []))
    _table("기준일 이전 누락 보충 — 원장에 있으나 기존 현황표에 없던 건",
           report.get("backfilled", []))


def save_result(df, out_path, 퇴사자_names=None, log_fn=print,
                report=None, ledger_balance=None):
    if 퇴사자_names is None:
        퇴사자_names = set()

    # ── 지급일자 통일 & 정렬: 현장명 → 지급일자 ──
    df = df.copy()
    df["지급일자"] = df["지급일자"].apply(_normalize_date)
    df["_sort_현장"] = df["현장명"].fillna("").apply(lambda x: _normalize(str(x)))
    df["_sort_날짜"] = df["지급일자"].apply(_date_to_sortkey)
    df = df.sort_values(["_sort_현장", "_sort_날짜"]).drop(columns=["_sort_현장", "_sort_날짜"])
    df = df.reset_index(drop=True)
    df["번호"] = range(1, len(df) + 1)

    # ── 직원 / 현장사무실 분리 ──
    # 지급자가 비었거나 '법인'이면 임대인·법인 명의 계약 → 현장사무실 시트
    mask_사무실 = df["지급자"].apply(
        lambda x: _normalize(str(x or "")) in _사무실_지급자)
    df_직원    = df[~mask_사무실].reset_index(drop=True)
    df_사무실  = df[mask_사무실].reset_index(drop=True)
    df_직원["번호"] = range(1, len(df_직원) + 1)
    df_사무실["번호"] = range(1, len(df_사무실) + 1)

    wb = Workbook()

    # 시트 1: 직원 숙소보증금
    ws1 = wb.active
    ws1.title = "직원숙소보증금"
    amt1 = _write_sheet(ws1, df_직원, 퇴사자_names, "직원숙소보증금")

    # 시트 2: 현장사무실 (지급자 없는 행)
    if len(df_사무실) > 0:
        ws2 = wb.create_sheet("현장사무실")
        amt2 = _write_sheet(ws2, df_사무실, set(), "현장사무실")
        log_fn(f"   현장사무실 시트: {len(df_사무실)}건 / {amt2:,}원")
    else:
        amt2 = 0

    # 시트 3: 원장 대조 + 매칭 실패
    if report is not None or ledger_balance is not None:
        ws3 = wb.create_sheet("원장대조")
        _write_check_sheet(ws3, report or {}, ledger_balance, amt1 + amt2, log_fn)

    wb.save(out_path)
    log_fn(f"   저장 완료: {os.path.basename(out_path)}")
    log_fn(f"   직원: {len(df_직원)}건 / {amt1:,}원")
    if ledger_balance is not None:
        diff = (amt1 + amt2) - ledger_balance
        log_fn(f"   원장 최종 잔액: {ledger_balance:,}원 / 차액: {diff:,}원")
    if 퇴사자_names:
        n = sum(1 for r in df_직원.itertuples()
                if _normalize(str(getattr(r, "지급자", "") or "")) in 퇴사자_names)
        if n:
            log_fn(f"   ⚠ 퇴사자 미반환: {n}건 (노란색 강조)")


# ──────────────────────────────────────────────
# 메인 처리 함수
# ──────────────────────────────────────────────

def process(ledger_paths, status_path, output_dir, hr_path=None, log_fn=print,
            start_date=None, backfill=True):
    """
    ledger_paths: 원장 파일 경로 문자열 또는 리스트
    start_date  : 'YYYY-MM-DD' — 이 날짜 이후(당일 포함) 거래만 반영.
                  기존 현황표가 특정 시점까지 이미 반영돼 있을 때 중복을 막는다.
    backfill    : 기준일 이전 원장에서 현황표에 빠진 미반환 건을 찾아 채운다.
    """
    if isinstance(ledger_paths, str):
        ledger_paths = [ledger_paths]
    ledger_paths = [p for p in ledger_paths if p and os.path.isfile(p)]

    log_fn("=" * 52)

    hr, 퇴사자_names = {}, set()
    if hr_path:
        log_fn("[0/4] 인사기록부 로딩...")
        hr, 퇴사자_names = parse_hr(hr_path, log_fn)

    log_fn(f"\n[1/4] 원장 파싱... ({len(ledger_paths)}개 파일)")
    txns = []
    ledger_balance = None
    latest_key = -1
    for p in ledger_paths:
        log_fn(f"   ▸ {os.path.basename(p)}")
        t, bal = parse_ledger(p, log_fn)
        txns.extend(t)
        # 가장 마지막 거래를 담은 원장의 잔액을 대조 기준으로 삼는다
        key = max([_date_to_sortkey(x["date_str"]) for x in t], default=-1)
        if bal is not None and key >= latest_key:
            latest_key, ledger_balance = key, bal

    # 날짜 순 정렬 (연도 섞여도 순서 보장)
    txns.sort(key=lambda t: _date_to_sortkey(t["date_str"]))
    log_fn(f"   전체 거래 합계: {len(txns)}건")

    past_txns = None
    if start_date:
        cut = _date_to_sortkey(start_date)
        if cut:
            past_txns = [t for t in txns if _date_to_sortkey(t["date_str"]) < cut]
            txns      = [t for t in txns if _date_to_sortkey(t["date_str"]) >= cut]
            log_fn(f"   기준일 {start_date} 이후 반영: {len(txns)}건 / "
                   f"이전 {len(past_txns)}건은 누락 점검에만 사용")
            if not backfill:
                past_txns = None
        else:
            log_fn(f"   ⚠ 기준일 '{start_date}' 형식을 알 수 없어 무시합니다")

    log_fn("\n[2/4] 기존 현황표 로딩...")
    if status_path and os.path.isfile(status_path):
        df = load_status(status_path, log_fn)
    else:
        cols = _HEADERS + ["지급자_norm"]
        df = pd.DataFrame(columns=cols)
        log_fn("   기존 현황표 없음 → 새로 생성")

    log_fn("\n[3/4] 업데이트 처리...")
    df, report = update_status(df, txns, hr=hr, log_fn=log_fn, past_txns=past_txns)

    log_fn("\n[4/4] 저장...")
    today = date.today().strftime("%Y.%m.%d")
    out_name = f"감리현장숙소계약현황_{today}기준.xlsx"
    out_path = os.path.join(output_dir, out_name)
    save_result(df, out_path, 퇴사자_names=퇴사자_names, log_fn=log_fn,
                report=report, ledger_balance=ledger_balance)

    log_fn("=" * 52)
    log_fn(f"완료: {out_name}")
    return out_path


# ──────────────────────────────────────────────
# GUI
# ──────────────────────────────────────────────

DROP_BG  = "#ffffff"
DROP_HOV = "#cfe2ff"


class DepositApp:
    def __init__(self, root):
        self.root = root
        root.title("숙소보증금 현황표 생성")
        root.geometry("700x700")
        root.resizable(True, True)

        self.ledger_paths = []   # 원장 파일 목록
        self.status_path  = tk.StringVar()
        self.hr_path      = tk.StringVar()
        self.output_dir   = tk.StringVar()
        self.start_date   = tk.StringVar()   # 기준일 (이 날짜 이후 거래만 반영)

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 12, "pady": 5}

        tk.Label(self.root, text="숙소보증금 현황표 생성",
                 font=("맑은 고딕", 15, "bold"), pady=6).pack()

        # ── 1. 원장 파일 (복수) ──
        f1 = ttk.LabelFrame(self.root, text="1. 임차보증금 계정별 원장  (.xlsx)  *필수  ─ 여러 연도 동시 추가 가능")
        f1.pack(fill="x", **pad)

        list_frame = tk.Frame(f1)
        list_frame.pack(fill="x", padx=8, pady=(6, 2))

        sb = tk.Scrollbar(list_frame, orient="vertical")
        self.ledger_listbox = tk.Listbox(
            list_frame, height=4, selectmode="extended",
            font=("맑은 고딕", 9), yscrollcommand=sb.set,
            activestyle="none",
        )
        sb.config(command=self.ledger_listbox.yview)
        self.ledger_listbox.pack(side="left", fill="x", expand=True)
        sb.pack(side="left", fill="y")

        btn_row = tk.Frame(f1)
        btn_row.pack(fill="x", padx=8, pady=(2, 6))
        ttk.Button(btn_row, text="파일 추가...", command=self.pick_ledger).pack(side="left", padx=(0, 4))
        ttk.Button(btn_row, text="선택 삭제",   command=self.remove_ledger).pack(side="left", padx=(0, 4))
        ttk.Button(btn_row, text="전체 삭제",   command=self.clear_ledger).pack(side="left")

        # 드래그앤드롭 — 리스트박스에 파일 드롭
        self._register_drop(self.ledger_listbox, self._on_drop_ledger)

        # ── 2. 기존 현황표 ──
        f2 = ttk.LabelFrame(self.root, text="2. 기존 현황표  (.xlsx)  ─ 없으면 원장 기준으로 신규 생성")
        f2.pack(fill="x", **pad)
        self._file_row(f2, self.status_path, self.pick_status, ext=".xlsx")

        # ── 3. 인사기록부 ──
        f3 = ttk.LabelFrame(self.root, text="3. 인사기록부  (.xls/.xlsx)  ─ 이름 자동 보정 (선택)")
        f3.pack(fill="x", **pad)
        self._file_row(f3, self.hr_path, self.pick_hr, ext=None)

        # ── 3-1. 기준일 ──
        f31 = ttk.LabelFrame(self.root, text="4. 기준일  ─ 기존 현황표가 반영된 마지막 날짜 (선택, 비우면 원장 전체 반영)")
        f31.pack(fill="x", **pad)
        r31 = tk.Frame(f31); r31.pack(fill="x", padx=8, pady=6)
        ttk.Entry(r31, textvariable=self.start_date, width=16).pack(side="left")
        tk.Label(r31, text="  예) 2025-10-29 입력 시 그 날짜부터의 거래만 반영",
                 font=("맑은 고딕", 8), fg="#666").pack(side="left")

        # ── 4. 저장 폴더 ──
        f4 = ttk.LabelFrame(self.root, text="5. 저장 폴더")
        f4.pack(fill="x", **pad)
        r4 = tk.Frame(f4); r4.pack(fill="x", padx=8, pady=6)
        ttk.Entry(r4, textvariable=self.output_dir).pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Button(r4, text="폴더 선택...", command=self.pick_output).pack(side="right")
        self._register_drop(r4, self._on_drop_dir)

        # ── 실행 ──
        self.run_btn = tk.Button(
            self.root, text="▶  현황표 생성",
            font=("맑은 고딕", 12, "bold"),
            bg="#1F497D", fg="white", relief="flat", cursor="hand2",
            command=self.run,
        )
        self.run_btn.pack(pady=8, ipadx=20, ipady=6)

        # ── 로그 ──
        lf = ttk.LabelFrame(self.root, text="실행 로그")
        lf.pack(fill="both", expand=True, **pad)
        self.log_box = scrolledtext.ScrolledText(lf, height=10, font=("Consolas", 9), wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=4, pady=4)

    # ── 원장 리스트 관리 ──

    def _add_ledger_path(self, path):
        path = path.strip().strip("{}")
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        if ext not in (".xlsx", ".xls"):
            return
        if not os.path.isfile(path):
            return
        if path in self.ledger_paths:
            return
        self.ledger_paths.append(path)
        self.ledger_listbox.insert("end", os.path.basename(path))
        if not self.output_dir.get():
            self.output_dir.set(os.path.dirname(path))

    def pick_ledger(self):
        paths = filedialog.askopenfilenames(
            parent=self.root,
            title="원장 파일 선택 (복수 선택 가능)",
            filetypes=[("Excel", "*.xlsx *.xls"), ("모든 파일", "*.*")],
        )
        for p in paths:
            self._add_ledger_path(p)

    def remove_ledger(self):
        for i in reversed(self.ledger_listbox.curselection()):
            self.ledger_listbox.delete(i)
            self.ledger_paths.pop(i)

    def clear_ledger(self):
        self.ledger_listbox.delete(0, "end")
        self.ledger_paths.clear()

    def _on_drop_ledger(self, event):
        self._hl(event.widget, False)
        for p in self._split(event.data):
            self._add_ledger_path(p)

    def _file_row(self, parent, var, cmd, ext):
        r = tk.Frame(parent); r.pack(fill="x", padx=8, pady=6)
        entry = ttk.Entry(r, textvariable=var)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Button(r, text="파일 선택...", command=cmd).pack(side="right")
        self._register_drop(r, lambda e, v=var, x=ext: self._on_drop_file(e, v, x))

    def _register_drop(self, widget, handler):
        if not DND_AVAILABLE:
            return
        try:
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", handler)
            widget.dnd_bind("<<DropEnter>>", lambda e: self._hl(e.widget, True))
            widget.dnd_bind("<<DropLeave>>", lambda e: self._hl(e.widget, False))
        except Exception:
            pass

    def _hl(self, w, on):
        try:
            w.config(bg=DROP_HOV if on else DROP_BG)
        except Exception:
            pass

    def _split(self, raw):
        try:
            return list(self.root.tk.splitlist(raw))
        except Exception:
            return [raw]

    def _on_drop_file(self, event, var, ext):
        self._hl(event.widget, False)
        for p in self._split(event.data):
            p = p.strip().strip("{}")
            if os.path.isfile(p):
                if ext is None or p.lower().endswith(ext):
                    var.set(p)
                    if not self.output_dir.get():
                        self.output_dir.set(os.path.dirname(p))
                    break

    def _on_drop_dir(self, event):
        self._hl(event.widget, False)
        for p in self._split(event.data):
            p = p.strip().strip("{}")
            d = p if os.path.isdir(p) else (os.path.dirname(p) if os.path.isfile(p) else None)
            if d:
                self.output_dir.set(d)
                break

    def pick_status(self):
        p = filedialog.askopenfilename(
            parent=self.root,
            title="기존 현황표 선택",
            filetypes=[("Excel", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if p:
            self.status_path.set(p)

    def pick_hr(self):
        p = filedialog.askopenfilename(
            parent=self.root,
            title="인사기록부 선택",
            filetypes=[("Excel", "*.xls *.xlsx"), ("모든 파일", "*.*")],
        )
        if p:
            self.hr_path.set(p)

    def pick_output(self):
        d = filedialog.askdirectory(parent=self.root, title="저장 폴더 선택")
        if d:
            self.output_dir.set(d)

    def _log(self, msg):
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.root.update_idletasks()

    def run(self):
        if not self.ledger_paths:
            messagebox.showwarning("입력 필요", "원장 파일을 하나 이상 추가하세요.")
            return
        if not self.output_dir.get():
            messagebox.showwarning("입력 필요", "저장 폴더를 선택하세요.")
            return

        self.log_box.delete("1.0", "end")
        self.run_btn.config(state="disabled", text="처리 중...")

        threading.Thread(
            target=self._worker,
            args=(
                list(self.ledger_paths),
                self.status_path.get() or None,
                self.output_dir.get(),
                self.hr_path.get() or None,
                self.start_date.get().strip() or None,
            ),
            daemon=True,
        ).start()

    def _worker(self, ledger, status, outdir, hr_path, start_date=None):
        try:
            result = process(ledger, status, outdir, hr_path=hr_path,
                             log_fn=self._log, start_date=start_date)
            self.root.after(0, lambda: messagebox.showinfo(
                "완료", f"현황표가 생성되었습니다.\n\n저장 위치:\n{result}"))
        except Exception as e:
            self._log(f"\n오류:\n{traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.root.after(0, lambda: self.run_btn.config(
                state="normal", text="▶  현황표 생성"))


def open_window(parent):
    win = tk.Toplevel(parent)
    DepositApp(win)
    return win


if __name__ == "__main__":
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except ImportError:
        root = tk.Tk()
    DepositApp(root)
    root.mainloop()
