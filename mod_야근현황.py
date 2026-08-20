"""
야근현황 자동 입력기 v5 — 허브용 모듈
원본: overtime_gui.py (최신)
변경: open_window(parent) 함수 추가
"""

import os
import re
import sys
import json
import threading
import traceback
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError as _e:
    import tkinter as _tk
    from tkinter import messagebox as _mb
    _root = _tk.Tk()
    _root.withdraw()
    _mb.showerror(
        "라이브러리 누락",
        f"필요한 라이브러리가 설치되어 있지 않습니다:\n\n{_e}\n\n"
        f"cmd에서: pip install pandas openpyxl xlrd tkinterdnd2",
    )
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False


IN_MODES = {'출근'}
OUT_MODES = {'퇴근'}


def extract_department(filename):
    name = os.path.splitext(os.path.basename(filename))[0]
    name = re.sub(r'_자동입력_\d{8}_\d{6}', '', name)
    name = re.sub(r'[_\s]?\d{4}년\s*\d{1,2}월', '', name)
    name = re.sub(r'[_\s]?\d{1,2}월', '', name)
    name = re.sub(r'[_\s]?\d{4}(?=[_\s]|$)', '', name)
    for kw in ('야근현황', '야근', '출퇴근'):
        name = name.replace(kw, '')
    name = name.strip('_').strip()
    return name or os.path.splitext(os.path.basename(filename))[0]


def detect_attendance_format(path):
    for header_row in (0, 1):
        try:
            df = pd.read_excel(path, sheet_name=0, header=header_row, nrows=1)
            cols = set(df.columns.astype(str))
            if {'인증일시', '인증모드', '이름'}.issubset(cols):
                return 'auth_log'
            if {'근무일자', '이름', '출근시간', '퇴근시간'}.issubset(cols):
                return 'daily_summary'
        except Exception:
            continue
    return None


def _time_to_hhmm(v):
    if pd.isna(v):
        return None
    if hasattr(v, 'hour') and hasattr(v, 'minute'):
        return v.hour * 100 + v.minute
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return None
    if ':' in s:
        parts = s.split(':')
        try:
            return int(parts[0]) * 100 + int(parts[1])
        except (ValueError, IndexError):
            return None
    return None


def _merge_entry(combined, key, entry):
    if key in combined:
        old = combined[key]
        if 'in' in entry and ('in' not in old or entry['in'] < old['in']):
            old['in'] = entry['in']
        if 'out' in entry and ('out' not in old or entry['out'] > old['out']):
            old['out'] = entry['out']
    else:
        combined[key] = dict(entry)


def load_attendance_auth_log(path):
    df = pd.read_excel(path, sheet_name=0, header=1)
    df = df.drop_duplicates(subset=['인증일시', '사원번호', '인증모드', '리더기 장소'])
    df['이름_정규'] = df['이름'].astype(str).str.replace(' ', '', regex=False)
    df['일시'] = pd.to_datetime(df['인증일시'])
    df['날짜'] = df['일시'].dt.date

    def classify(mode):
        if mode in IN_MODES: return 'IN'
        if mode in OUT_MODES: return 'OUT'
        return None
    df['구분'] = df['인증모드'].apply(classify)
    df = df[df['구분'].notna()]

    result = {}
    for (name, d), g in df.groupby(['이름_정규', '날짜']):
        ins = g[g['구분'] == 'IN']['일시']
        outs = g[g['구분'] == 'OUT']['일시']
        entry = {}
        if len(ins) > 0:
            t = ins.min()
            entry['in'] = t.hour * 100 + t.minute
        if len(outs) > 0:
            t = outs.max()
            entry['out'] = t.hour * 100 + t.minute
        if entry:
            result[(name, d)] = entry

    dups = {}
    for name, g in df.groupby('이름_정규'):
        ids = g['사원번호'].dropna().unique()
        if len(ids) > 1:
            dups[name] = list(ids)

    return result, dups, len(df), df['이름_정규'].nunique()


def load_attendance_daily_summary(path):
    df = pd.read_excel(path, sheet_name=0, header=0)
    df = df[df['이름'].notna()].copy()
    df['이름_정규'] = df['이름'].astype(str).str.replace(' ', '', regex=False)
    df['날짜'] = pd.to_datetime(df['근무일자']).dt.date

    result = {}
    nrec = 0
    for _, row in df.iterrows():
        name = row['이름_정규']
        d = row['날짜']
        in_t = _time_to_hhmm(row['출근시간'])
        out_t = _time_to_hhmm(row['퇴근시간'])
        if in_t is None and out_t is None:
            continue
        entry = {}
        if in_t is not None: entry['in'] = in_t
        if out_t is not None: entry['out'] = out_t
        _merge_entry(result, (name, d), entry)
        nrec += 1

    return result, {}, nrec, df['이름_정규'].nunique()


def load_all_attendance(paths, log_fn=print):
    combined = {}
    all_dups = {}
    total_records = 0
    all_names = set()

    for p in paths:
        fmt = detect_attendance_format(p)
        if fmt is None:
            log_fn(f"   ⚠ '{os.path.basename(p)}': 알 수 없는 형식 — 건너뜀")
            continue

        try:
            if fmt == 'auth_log':
                daily, dups, nrec, npeople = load_attendance_auth_log(p)
                log_fn(f"   • {os.path.basename(p)}  [인증로그]  레코드 {nrec}건, 인원 {npeople}명")
            else:
                daily, dups, nrec, npeople = load_attendance_daily_summary(p)
                log_fn(f"   • {os.path.basename(p)}  [일별요약]  레코드 {nrec}건, 인원 {npeople}명")
        except Exception as e:
            log_fn(f"   ❌ '{os.path.basename(p)}' 로딩 실패: {e}")
            continue

        total_records += nrec
        for key, entry in daily.items():
            all_names.add(key[0])
            _merge_entry(combined, key, entry)
        all_dups.update(dups)

    return combined, all_dups, total_records, len(all_names)


def calc_period(year, month):
    if month == 1:
        start = date(year - 1, 12, 21)
    else:
        start = date(year, month - 1, 21)
    end = date(year, month, 20)
    return start, end


def parse_user_date(s):
    if not s or not s.strip():
        return None
    s = s.strip()
    digits = re.sub(r'[^\d]', '', s)
    if digits == s.replace(' ', ''):
        today = date.today()
        try:
            if len(digits) == 8:
                return date(int(digits[0:4]), int(digits[4:6]), int(digits[6:8]))
            elif len(digits) == 6:
                return date(2000 + int(digits[0:2]), int(digits[2:4]), int(digits[4:6]))
            elif len(digits) == 4:
                return date(today.year, int(digits[0:2]), int(digits[2:4]))
        except ValueError as e:
            raise ValueError(f"잘못된 날짜: '{s}' ({e})")
        raise ValueError(f"날짜 형식 오류: '{s}'")
    s = s.replace('.', '-').replace('/', '-')
    parts = s.split('-')
    if len(parts) != 3:
        raise ValueError(f"날짜 형식 오류: '{s}'")
    try:
        y, m, d_ = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
        return date(y, m, d_)
    except ValueError as e:
        raise ValueError(f"잘못된 날짜: '{s}' ({e})")


def detect_form_layout(ws):
    year = ws.cell(2, 2).value
    if not isinstance(year, int) or not (2000 <= year <= 2100):
        return None, None
    month_col = None
    for col in range(3, 20):
        v = ws.cell(2, col).value
        if isinstance(v, int) and 1 <= v <= 12:
            month_col = col
            break
    date_start_col = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(3, col).value
        if isinstance(v, str) and '시작일' in v:
            date_start_col = col
            break
        if isinstance(v, datetime):
            date_start_col = col
            break
        if isinstance(v, str) and v.startswith('='):
            date_start_col = col
            break
    return month_col, date_start_col


def is_overtime_sheet(ws):
    try:
        if ws.cell(3, 1).value != '구분':
            return False
        if ws.cell(3, 3).value != '성명':
            return False
        month_col, date_start_col = detect_form_layout(ws)
        return month_col is not None and date_start_col is not None
    except Exception:
        return False


def _is_formula(cell_value):
    return isinstance(cell_value, str) and cell_value.startswith('=')


def _extract_formula_templates(ws, date_start_col, num_days):
    """템플릿 시트의 성명행 야근 수식을 행번호 독립 템플릿으로 뽑는다.

    데이터행 참조만 '{ROW}' 로 치환한다. 예전에는 수식 전체를
    str.replace(str(data_row), '{ROW}') 로 바꿨는데, 그러면 수식 안의
    숫자 리터럴까지 함께 잡혔다. 예를 들어 data_row 가 10 이면
    INT(F10/100) 이 INT(F{ROW}/{ROW}0) 이 되고, 이 템플릿이 fallback 으로
    다른 행에 적용되면 /100 이 /120 처럼 깨진 수식이 만들어졌다.

    _shift_formula_rows 는 셀 참조의 행번호만 건드리므로(절대행 $6 ·
    명명범위 · 함수명 제외) 리터럴이 오염되지 않는다.
    """
    per_row = {}
    fallback = {}
    for row in range(7, 80, 2):
        if ws.cell(row, 3).value is None:
            continue
        data_row = row + 1
        row_tmpl = {}
        for d_off in range(num_days):
            col = date_start_col + 2 * d_off
            v = ws.cell(row, col).value
            if _is_formula(v):
                row_tmpl[d_off] = _shift_formula_rows(v, {data_row: '{ROW}'})
        if row_tmpl:
            per_row[row] = row_tmpl
            if not fallback:
                fallback = row_tmpl
    return per_row, fallback


def _apply_employee_rule(formula, name, employee_rules):
    if not employee_rules or not name:
        return formula
    rule = employee_rules.get(name.replace(' ', ''))
    if not rule:
        return formula
    ot = rule.get('야근시작')
    if not ot:
        return formula
    try:
        h, m = map(int, ot.split(':'))
        return formula.replace('야근시작', f'TIME({h},{m},0)')
    except Exception:
        return formula


def _restore_name_row_formulas(ws, date_start_col, num_days, per_row, fallback,
                               employee_rules=None):
    if not per_row and not fallback:
        return 0
    restored = 0
    for row in range(7, 80, 2):
        name = ws.cell(row, 3).value
        if name is None:
            continue
        data_row = row + 1
        templates = per_row.get(row, fallback)
        for d_off in range(num_days):
            col = date_start_col + 2 * d_off
            cell = ws.cell(row, col)
            if cell.value is None and d_off in templates:
                formula = templates[d_off].replace('{ROW}', str(data_row))
                formula = _apply_employee_rule(formula, str(name), employee_rules or {})
                cell.value = formula
                restored += 1
    return restored


def clear_period_cells(ws, date_start_col, form_start, eff_start, eff_end,
                       protect_text=False):
    range_start = (eff_start - form_start).days
    range_end = (eff_end - form_start).days
    for row in range(7, 80):
        is_name_row = (row % 2 == 1)
        for d_off in range(range_start, range_end + 1):
            col_in = date_start_col + 2 * d_off
            if is_name_row:
                cell = ws.cell(row, col_in)
                if not _is_formula(cell.value):
                    cell.value = None
            else:
                for col in (col_in, col_in + 1):
                    cell = ws.cell(row, col)
                    if protect_text and isinstance(cell.value, str):
                        continue
                    cell.value = None


# ── 출퇴근 데이터 초기화 ──────────────────────────────────────────────
# 양식의 일자 칸은 하루당 2열(출근/퇴근)이며, 책정기간이 30일인 달에도
# 31일치 칸이 만들어져 있다. 초기화는 책정기간이 아니라 '시트에 실재하는
# 일자 칸 전체'를 대상으로 해야 지난달 31일차 잔여 데이터까지 지워진다.
def count_day_slots(ws, date_start_col, max_slots=40):
    """3행의 날짜 셀만 2열 간격으로 센다.

    일자 칸 오른쪽에는 '평일야근', '보상휴가' 같은 요약 열이 이어진다.
    그 헤더는 일반 텍스트이므로, 날짜(datetime) 또는 날짜 수식만 인정해야
    요약 열을 일자 칸으로 잘못 세지 않는다.
    """
    n = 0
    for i in range(max_slots):
        v = ws.cell(3, date_start_col + 2 * i).value
        if isinstance(v, datetime) or _is_formula(v):
            n += 1
            continue
        break
    return n


# 수식의 셀 참조 행번호만 옮긴다. 절대행($6)·명명범위(야근시작)·
# 함수명(LOG10 처럼 뒤에 '(' 가 오는 것)은 건드리지 않는다.
_CELL_REF_RE = re.compile(r'(?<![A-Za-z0-9_.$!])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![\d(])')


def _shift_formula_rows(formula, row_map):
    def repl(m):
        d1, col, d2, row = m.groups()
        if d2 == '$':
            return m.group(0)
        return f"{d1}{col}{d2}{row_map.get(int(row), int(row))}"
    return _CELL_REF_RE.sub(repl, formula)


def find_daycell_formula_donor(ws, date_start_col, num_slots):
    """성명행 중 날짜칸 수식이 가장 온전한 행을 골라 (행번호, {칸번호: 수식})."""
    best_row, best_tmpl = None, {}
    for row in range(7, 80, 2):
        tmpl = {}
        for i in range(num_slots):
            v = ws.cell(row, date_start_col + 2 * i).value
            if _is_formula(v):
                tmpl[i] = v
        if len(tmpl) > len(best_tmpl):
            best_row, best_tmpl = row, tmpl
    return best_row, best_tmpl


def restore_daycell_formulas(ws, date_start_col, num_slots, employee_rules=None):
    """같은 시트의 온전한 행에서 야근 계산 수식을 복사해 빈 성명행에 채운다.

    별도 템플릿 파일 없이 시트 내부만으로 복원하므로, 수기 작업 중
    수식이 값으로 덮인 행을 그 자리에서 되살릴 수 있다.
    """
    donor_row, tmpl = find_daycell_formula_donor(ws, date_start_col, num_slots)
    if not tmpl:
        return 0, None
    donor_data = donor_row + 1
    restored = 0
    for row in range(7, 80, 2):
        if row == donor_row:
            continue
        name = ws.cell(row, 3).value
        if not name:
            continue
        data_row = row + 1
        row_map = {donor_row: row, donor_data: data_row}
        for i, f in tmpl.items():
            cell = ws.cell(row, date_start_col + 2 * i)
            if _is_formula(cell.value):
                continue
            new_f = _shift_formula_rows(f, row_map)
            new_f = _apply_employee_rule(new_f, str(name), employee_rules or {})
            cell.value = new_f
            restored += 1
    return restored, donor_row


def clear_attendance_cells(ws, date_start_col, num_slots, clear_reason=False):
    """데이터행의 출근/퇴근 값을 비운다. 수식은 모두 보존한다.

    성명행에 남은 '수식이 아닌 값'(수식이 값으로 덮인 흔적)도 함께 지운다.
    이 값을 남기면 출퇴근을 지운 뒤에도 옛 야근시간이 표시돼 버린다.
    """
    cleared_data = 0
    cleared_stale = 0
    for row in range(7, 80):
        is_name_row = (row % 2 == 1)
        for i in range(num_slots):
            col_in = date_start_col + 2 * i
            if is_name_row:
                cell = ws.cell(row, col_in)
                if cell.value is not None and not _is_formula(cell.value):
                    cell.value = None
                    cleared_stale += 1
            else:
                for col in (col_in, col_in + 1):
                    cell = ws.cell(row, col)
                    if cell.value is None or _is_formula(cell.value):
                        continue
                    if not clear_reason and isinstance(cell.value, str):
                        continue
                    cell.value = None
                    cleared_data += 1
    return cleared_data, cleared_stale


def process_clear_file(path, output_dir, log_fn, clear_reason=False,
                       do_restore=True, overwrite=False, employee_rules=None):
    log_fn(f"\n▶ 초기화: {os.path.basename(path)}")
    wb = load_workbook(path)
    processed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not is_overtime_sheet(ws):
            log_fn(f"    [{sheet_name}] 스킵 (양식 아님)")
            continue

        _, dsc = detect_form_layout(ws)
        num_slots = count_day_slots(ws, dsc)
        if num_slots == 0:
            log_fn(f"    [{sheet_name}] ⚠ 일자 칸을 찾지 못함 — 건너뜀")
            continue

        cd, cs = clear_attendance_cells(ws, dsc, num_slots, clear_reason)
        log_fn(f"    [{sheet_name}] 일자 칸 {num_slots}개 · 출퇴근 {cd}셀 삭제"
               + ("" if clear_reason else "  [사유 텍스트 보호]"))
        if cs:
            log_fn(f"    [{sheet_name}]   └ 성명행에 값으로 덮여있던 야근시간 {cs}셀 제거")

        if do_restore:
            n, donor = restore_daycell_formulas(ws, dsc, num_slots, employee_rules)
            if n:
                log_fn(f"    [{sheet_name}]   └ 야근 계산 수식 복원 {n}셀 (기준 행 {donor})")
            elif cs:
                log_fn(f"    [{sheet_name}]   ⚠ 수식 복원 실패 — 시트에 온전한 수식 행이 없음")

        processed += 1

    if processed == 0:
        log_fn("    ⚠ 처리 가능한 시트 없음")
        return None

    if overwrite:
        out_path = path
    else:
        base, ext = os.path.splitext(os.path.basename(path))
        out_path = os.path.join(output_dir, f"{base}_초기화{ext}")
    wb.save(out_path)
    log_fn(f"    💾 저장: {os.path.basename(out_path)}")
    return out_path


def fill_sheet(ws, daily_map, log_fn, label, user_start=None, user_end=None,
               protect_text=False):
    month_col, date_start_col = detect_form_layout(ws)
    year = ws.cell(2, 2).value
    month = ws.cell(2, month_col).value
    form_start, form_end = calc_period(year, month)

    eff_start = max(form_start, user_start) if user_start else form_start
    eff_end = min(form_end, user_end) if user_end else form_end

    if eff_start > eff_end:
        log_fn(f"    [{label}] ⚠ 지정 기간이 양식 책정기간({form_start}~{form_end})과 겹치지 않음 — 건너뜀")
        return 0, []

    log_fn(f"    [{label}] 처리 범위: {eff_start} ~ {eff_end}")

    num_days = (eff_end - eff_start).days + 1
    clear_period_cells(ws, date_start_col, form_start, eff_start, eff_end,
                       protect_text=protect_text)
    log_fn(f"    [{label}]   └ 기존 데이터 초기화 완료 ({num_days}일분)"
           + (" [사유 보호]" if protect_text else ""))

    employees = []
    for row in range(7, 80, 2):
        nm = ws.cell(row, 3).value
        if nm:
            employees.append((row + 1, str(nm).replace(' ', '')))

    range_start_off = (eff_start - form_start).days
    range_end_off = (eff_end - form_start).days

    filled = 0
    matched, unmatched = [], []
    for data_row, name in employees:
        any_data = False
        for d_off in range(range_start_off, range_end_off + 1):
            d = form_start + timedelta(days=d_off)
            entry = daily_map.get((name, d))
            if not entry:
                continue
            any_data = True
            col_in  = date_start_col + 2 * d_off
            col_out = date_start_col + 1 + 2 * d_off
            if 'in' in entry:
                cell = ws.cell(data_row, col_in)
                if not (protect_text and isinstance(cell.value, str)):
                    cell.value = entry['in']
                    filled += 1
            if 'out' in entry:
                cell = ws.cell(data_row, col_out)
                if not (protect_text and isinstance(cell.value, str)):
                    cell.value = entry['out']
                    filled += 1
        (matched if any_data else unmatched).append(name)

    log_fn(f"    [{label}]   └ 출퇴근 데이터 입력: {filled}셀 / 직원 {len(employees)}명 중 {len(matched)}명 매칭")
    if unmatched:
        log_fn(f"        ⚠ 출퇴근 기록 미매칭 직원: {', '.join(unmatched)}")
    return len(matched), unmatched


def _get_template_ws(template_path):
    if not template_path or not os.path.isfile(template_path):
        return None, None
    try:
        wb = load_workbook(template_path)
        for sn in wb.sheetnames:
            ws = wb[sn]
            if is_overtime_sheet(ws):
                _, dsc = detect_form_layout(ws)
                return ws, dsc
    except Exception:
        pass
    return None, None


def process_modify_file(target_path, daily_map, output_dir, log_fn,
                        do_attendance, do_restore, do_protect,
                        template_normal, template_secha,
                        user_start=None, user_end=None,
                        employee_rules=None):
    log_fn(f"\n▶ 수정: {os.path.basename(target_path)}")
    wb = load_workbook(target_path)
    processed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not is_overtime_sheet(ws):
            log_fn(f"    [{sheet_name}] 스킵 (양식 아님)")
            continue

        month_col, date_start_col = detect_form_layout(ws)
        year  = ws.cell(2, 2).value
        month = ws.cell(2, month_col).value
        form_start, form_end = calc_period(year, month)
        num_days = (form_end - form_start).days + 1

        tmpl_path = template_secha if month_col == 6 else template_normal

        if do_restore:
            tmpl_ws, tmpl_dsc = _get_template_ws(tmpl_path)
            if tmpl_ws and tmpl_dsc:
                per_row, fallback = _extract_formula_templates(tmpl_ws, tmpl_dsc, num_days)
                restored = _restore_name_row_formulas(ws, date_start_col, num_days, per_row, fallback,
                                                      employee_rules)
                log_fn(f"    [{sheet_name}] 수식 복원: {restored}셀")
            else:
                log_fn(f"    [{sheet_name}] ⚠ 수식 복원 — 템플릿 없음 또는 감지 실패")

        if do_attendance and daily_map:
            eff_start = max(form_start, user_start) if user_start else form_start
            eff_end   = min(form_end,   user_end)   if user_end   else form_end
            if eff_start <= eff_end:
                fill_sheet(ws, daily_map, log_fn, sheet_name,
                           user_start, user_end, protect_text=do_protect)
            else:
                log_fn(f"    [{sheet_name}] ⚠ 지정 기간이 책정기간과 겹치지 않음")

        processed += 1

    if processed == 0:
        log_fn(f"    ⚠ 처리 가능한 시트 없음")
        return None

    dept = extract_department(target_path)
    out_year = out_month = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        mc, _ = detect_form_layout(ws)
        if mc:
            out_year  = ws.cell(2, 2).value
            out_month = ws.cell(2, mc).value
            break
    out_name = (f"{out_year}년{out_month}월_{dept}.xlsx"
                if out_year and out_month else os.path.basename(target_path))
    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)
    log_fn(f"    💾 저장: {os.path.basename(out_path)}")
    return out_path


def process_form_file(form_path, daily_map, output_dir, log_fn,
                       user_start=None, user_end=None,
                       target_year=None, target_month=None,
                       employee_rules=None):
    log_fn(f"\n▶ 양식: {os.path.basename(form_path)}")
    wb = load_workbook(form_path)
    processed = 0
    new_mode = target_year is not None and target_month is not None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not is_overtime_sheet(ws):
            log_fn(f"    [{sheet_name}] 스킵 (양식 아님)")
            continue

        if new_mode:
            month_col, date_start_col = detect_form_layout(ws)
            if month_col is not None:
                old_y = ws.cell(2, 2).value
                old_m = ws.cell(2, month_col).value
                ws.cell(2, 2).value = target_year
                ws.cell(2, month_col).value = target_month
                log_fn(f"    [{sheet_name}] {old_y}년 {old_m}월 → {target_year}년 {target_month}월로 변경")
                full_start, full_end = calc_period(target_year, target_month)
                num_days = (full_end - full_start).days + 1
                per_row, fallback = _extract_formula_templates(ws, date_start_col, num_days)
                clear_period_cells(ws, date_start_col, full_start, full_start, full_end)
                log_fn(f"    [{sheet_name}] 전체 기간 완전 초기화 ({full_start} ~ {full_end})")
                restored = _restore_name_row_formulas(ws, date_start_col, num_days, per_row, fallback,
                                                      employee_rules)
                if restored:
                    log_fn(f"    [{sheet_name}] 야근 계산 수식 복원: {restored}셀")

        fill_sheet(ws, daily_map, log_fn, sheet_name, user_start, user_end)
        processed += 1

    if processed == 0:
        log_fn(f"    ⚠ 처리 가능한 시트 없음")
        return None

    dept = extract_department(form_path)
    if new_mode:
        out_name = f"{target_year}년{target_month}월_{dept}.xlsx"
    else:
        out_year = out_month = None
        for sn in wb.sheetnames:
            ws = wb[sn]
            mc, _ = detect_form_layout(ws)
            if mc is not None:
                out_year = ws.cell(2, 2).value
                out_month = ws.cell(2, mc).value
                break
        if out_year and out_month:
            out_name = f"{out_year}년{out_month}월_{dept}.xlsx"
        else:
            out_name = f"{dept}_자동입력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)
    log_fn(f"    💾 저장: {os.path.basename(out_path)}")
    return out_path


def process_all(attendance_paths, form_paths, output_dir, log_fn,
                user_start=None, user_end=None,
                target_year=None, target_month=None,
                employee_rules=None):
    new_mode = target_year is not None and target_month is not None

    log_fn(f"[1/2] 출퇴근 파일 {len(attendance_paths)}개 로딩 (형식 자동 감지)...")
    daily_map, dups, total_rec, npeople = load_all_attendance(attendance_paths, log_fn)
    log_fn(f"   ━━ 합계: 유효 레코드 {total_rec:,}건, 총 인원 {npeople}명")
    log_fn(f"   ━━ (이름×날짜) 조합: {len(daily_map):,}건")

    if new_mode:
        log_fn(f"\n[모드] 새 양식 생성 — {target_year}년 {target_month}월")
        if not (user_start or user_end):
            user_start, user_end = calc_period(target_year, target_month)
            log_fn(f"[처리 범위 자동] {user_start} ~ {user_end}")
        else:
            log_fn(f"[처리 범위 지정] {user_start} ~ {user_end}")
    elif user_start or user_end:
        log_fn(f"\n[처리 범위 지정] {user_start or '(처음부터)'} ~ {user_end or '(끝까지)'}")

    if dups:
        log_fn(f"\n[⚠ 주의] 동명이인 발견:")
        for name, ids in dups.items():
            log_fn(f"   • '{name}' 사원번호 {ids}")

    log_fn(f"\n[2/2] 양식 파일 {len(form_paths)}개 처리...")
    saved = []
    for fp in form_paths:
        try:
            r = process_form_file(fp, daily_map, output_dir, log_fn,
                                   user_start, user_end, target_year, target_month,
                                   employee_rules)
            if r: saved.append(r)
        except Exception as e:
            log_fn(f"    ❌ 처리 실패: {e}")

    log_fn(f"\n{'='*50}")
    log_fn(f"✅ 전체 완료 — 생성 파일 {len(saved)}개")
    for f in saved:
        log_fn(f"   {os.path.basename(f)}")


def parse_dropped_paths(raw, root):
    try:
        return list(root.tk.splitlist(raw))
    except Exception:
        return [raw]


def expand_paths(paths, extensions):
    result = []
    for p in paths:
        p = p.strip().strip('{}').strip()
        if not p:
            continue
        if os.path.isdir(p):
            for f in sorted(os.listdir(p)):
                fp = os.path.join(p, f)
                if (os.path.isfile(fp)
                        and any(f.lower().endswith(ext) for ext in extensions)
                        and not f.startswith('~')):
                    result.append(fp)
        elif os.path.isfile(p) and any(p.lower().endswith(ext) for ext in extensions):
            result.append(p)
    return result


DROP_BG_NORMAL = '#ffffff'
DROP_BG_HOVER = '#cfe2ff'

_CONFIG_PATH = os.path.join(
    os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__),
    'overtime_config.json'
)


def _load_config():
    try:
        with open(_CONFIG_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_config(data):
    try:
        cfg = _load_config()
        cfg.update(data)
        with open(_CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


class OvertimeInputApp:
    def __init__(self, root):
        self.root = root
        root.title("야근현황 자동 입력기 v5")
        root.geometry("820x920")

        self.attendance_paths = []
        self.form_paths = []
        self.output_dir      = tk.StringVar()
        self.user_start      = tk.StringVar()
        self.user_end        = tk.StringVar()
        self.mode            = tk.StringVar(value='new')
        self.target_year     = tk.StringVar()
        self.target_month    = tk.StringVar()
        self.opt_attendance  = tk.BooleanVar(value=True)
        self.opt_restore     = tk.BooleanVar(value=True)
        self.opt_protect     = tk.BooleanVar(value=True)
        self.clr_reason      = tk.BooleanVar(value=False)
        self.clr_restore     = tk.BooleanVar(value=True)
        self.clr_overwrite   = tk.BooleanVar(value=False)
        self.tmpl_normal     = tk.StringVar()
        self.tmpl_secha      = tk.StringVar()
        self.employee_rules  = {}

        self._build_ui()
        self._on_mode_change()

        cfg = _load_config()
        if cfg.get('output_dir') and os.path.isdir(cfg['output_dir']):
            self.output_dir.set(cfg['output_dir'])
        if cfg.get('tmpl_normal'):
            self.tmpl_normal.set(cfg['tmpl_normal'])
        if cfg.get('tmpl_secha'):
            self.tmpl_secha.set(cfg['tmpl_secha'])
        rules = cfg.get('employee_rules', {})
        if rules:
            self.employee_rules = rules
            self._refresh_rule_list()

    def _build_ui(self):
        pad = {'padx': 12, 'pady': 5}

        outer = tk.Frame(self.root)
        outer.pack(fill='both', expand=True)

        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        inner = tk.Frame(canvas)
        canvas_win = canvas.create_window((0, 0), window=inner, anchor='nw')

        def _on_frame_resize(e):
            canvas.configure(scrollregion=canvas.bbox('all'))
            canvas.itemconfig(canvas_win, width=canvas.winfo_width())

        def _on_canvas_resize(e):
            canvas.itemconfig(canvas_win, width=e.width)

        inner.bind('<Configure>', _on_frame_resize)
        canvas.bind('<Configure>', _on_canvas_resize)

        def _on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind_all('<MouseWheel>', _on_mousewheel)

        # pack 대상을 inner로 임시 전환
        self.root = inner

        tk.Label(self.root, text="야근현황 자동 입력기",
                 font=('맑은 고딕', 16, 'bold'), pady=6).pack()
        hint = ("파일/폴더를 끌어다 놓거나 버튼으로 추가하세요"
                if DND_AVAILABLE else
                "버튼으로 파일을 추가하세요  (※ tkinterdnd2 미설치 — 드래그앤드롭 비활성)")
        tk.Label(self.root, text=hint, font=('맑은 고딕', 9), fg='gray').pack()

        f0 = ttk.LabelFrame(self.root, text="0. 실행 모드")
        f0.pack(fill='x', **pad)
        r0 = tk.Frame(f0); r0.pack(fill='x', padx=8, pady=6)
        ttk.Radiobutton(r0, text="새 양식 생성  — 전달 양식에서 명단 추출 → 지정 월 새 파일 생성",
                        variable=self.mode, value='new',
                        command=self._on_mode_change).pack(anchor='w')
        ttk.Radiobutton(r0, text="수정  — 기존 야근현황 파일 수정 (출퇴근 추가 / 수식 복원 / 사유 보호)",
                        variable=self.mode, value='modify',
                        command=self._on_mode_change).pack(anchor='w')
        ttk.Radiobutton(r0, text="초기화  — 출퇴근 데이터만 한 번에 비우기 (수식·서식·명단 유지)",
                        variable=self.mode, value='clear',
                        command=self._on_mode_change).pack(anchor='w')

        self._f1 = ttk.LabelFrame(self.root, text="1. 출퇴근 기록부  (.xls / .xlsx)")
        self._f1.pack(fill='x', **pad)
        self.att_list = tk.Listbox(self._f1, height=3, bg=DROP_BG_NORMAL)
        self.att_list.pack(fill='x', padx=8, pady=4)
        self._register_drop_target(self.att_list, self._on_drop_attendance)
        r1 = tk.Frame(self._f1); r1.pack(fill='x', padx=8, pady=4)
        ttk.Button(r1, text="파일 추가...", command=self.add_attendance).pack(side='left', padx=2)
        ttk.Button(r1, text="선택 삭제", command=self.del_attendance).pack(side='left', padx=2)
        ttk.Button(r1, text="모두 지우기", command=self.clear_attendance).pack(side='left', padx=2)

        self._f2 = ttk.LabelFrame(self.root, text="2. 양식 파일  (.xlsx)")
        self._f2.pack(fill='x', **pad)
        self.form_list = tk.Listbox(self._f2, height=4, bg=DROP_BG_NORMAL)
        self.form_list.pack(fill='x', padx=8, pady=4)
        self._register_drop_target(self.form_list, self._on_drop_forms)
        r2 = tk.Frame(self._f2); r2.pack(fill='x', padx=8, pady=4)
        ttk.Button(r2, text="파일 추가...", command=self.add_forms).pack(side='left', padx=2)
        ttk.Button(r2, text="폴더에서 모두 추가...", command=self.add_forms_folder).pack(side='left', padx=2)
        ttk.Button(r2, text="선택 삭제", command=self.del_forms).pack(side='left', padx=2)
        ttk.Button(r2, text="모두 지우기", command=self.clear_forms).pack(side='left', padx=2)

        self.f3a = ttk.LabelFrame(self.root, text="3. 생성할 연월")
        r3a = tk.Frame(self.f3a); r3a.pack(fill='x', padx=8, pady=6)
        tk.Label(r3a, text="연도:").pack(side='left')
        ttk.Entry(r3a, textvariable=self.target_year, width=8).pack(side='left', padx=(4, 16))
        tk.Label(r3a, text="월:").pack(side='left')
        ttk.Spinbox(r3a, textvariable=self.target_month, from_=1, to=12, width=5).pack(side='left')
        tk.Label(self.f3a,
                 text="※ 처리 범위는 지정 월 책정기간(전월21일~당월20일) 자동 설정",
                 font=('맑은 고딕', 8), fg='gray').pack(anchor='w', padx=12, pady=(0, 4))

        self.f3b = ttk.LabelFrame(self.root, text="3. 수정 옵션")
        rb = tk.Frame(self.f3b); rb.pack(fill='x', padx=8, pady=6)
        self._cb_att = ttk.Checkbutton(rb, text="출퇴근 시간 추가/수정  (출퇴근 기록부 필요)",
                                       variable=self.opt_attendance,
                                       command=self._on_opt_change)
        self._cb_att.pack(anchor='w')
        ttk.Checkbutton(rb, text="야근 계산 수식 복원  (아래 설정의 템플릿 사용)",
                        variable=self.opt_restore).pack(anchor='w')
        ttk.Checkbutton(rb, text="사유 입력 보호  (휴가·출장 등 텍스트 셀 건드리지 않음)",
                        variable=self.opt_protect).pack(anchor='w')

        self.f3c = ttk.LabelFrame(self.root, text="3. 초기화 옵션")
        rc = tk.Frame(self.f3c); rc.pack(fill='x', padx=8, pady=6)
        ttk.Checkbutton(rc, text="야근 계산 수식 자동 복원  (같은 시트의 온전한 행에서 복사)",
                        variable=self.clr_restore).pack(anchor='w')
        ttk.Checkbutton(rc, text="사유 텍스트도 삭제  (휴가·출장 등 한글 입력까지 비움)",
                        variable=self.clr_reason).pack(anchor='w')
        ttk.Checkbutton(rc, text="원본 파일에 덮어쓰기  (해제 시 '_초기화' 사본 저장)",
                        variable=self.clr_overwrite,
                        command=self._on_clr_change).pack(anchor='w')
        tk.Label(self.f3c,
                 text="※ 출퇴근 시간만 비웁니다. 직원 명단·설정·공휴일·수식·서식은 그대로 유지됩니다.",
                 font=('맑은 고딕', 8), fg='gray').pack(anchor='w', padx=12, pady=(0, 4))

        self.f4 = ttk.LabelFrame(self.root, text="4. 처리 기간 (선택)  — 비워두면 책정기간 전체 자동 사용")
        self.f4.pack(fill='x', **pad)
        r4 = tk.Frame(self.f4); r4.pack(fill='x', padx=8, pady=6)
        tk.Label(r4, text="시작일:").pack(side='left')
        ttk.Entry(r4, textvariable=self.user_start, width=14).pack(side='left', padx=(4, 12))
        tk.Label(r4, text="종료일:").pack(side='left')
        ttk.Entry(r4, textvariable=self.user_end, width=14).pack(side='left', padx=(4, 12))
        ttk.Button(r4, text="지우기", command=self.clear_dates).pack(side='left', padx=2)
        tk.Label(self.f4, text="형식: 20260521 · 260521 · 0521 · 2026-05-21",
                 font=('맑은 고딕', 8), fg='gray').pack(anchor='w', padx=12, pady=(0, 4))

        f5 = ttk.LabelFrame(self.root, text="5. 결과 저장 폴더")
        self.f5 = f5
        f5.pack(fill='x', **pad)
        r5 = tk.Frame(f5); r5.pack(fill='x', padx=8, pady=6)
        ttk.Entry(r5, textvariable=self.output_dir).pack(side='left', fill='x', expand=True, padx=(0, 8))
        ttk.Button(r5, text="폴더 선택...", command=self.pick_output_dir).pack(side='right')
        self._register_drop_target(r5, self._on_drop_outdir)

        self.run_btn = tk.Button(
            self.root, text="▶  일 괄  실 행",
            font=('맑은 고딕', 12, 'bold'),
            bg='#2563eb', fg='white', relief='flat', cursor='hand2',
            command=self.run,
        )
        self.run_btn.pack(pady=8, ipadx=20, ipady=6)

        fe = ttk.LabelFrame(self.root, text="직원 특이사항  (야근 시작 시간 개별 지정 — 저장됨)")
        fe.pack(fill='x', **pad)
        fe_top = tk.Frame(fe); fe_top.pack(fill='x', padx=8, pady=(6, 2))
        tk.Label(fe_top, text="이름", width=10, anchor='w', font=('맑은 고딕', 9, 'bold')).pack(side='left')
        tk.Label(fe_top, text="야근 시작 시간", width=14, anchor='w', font=('맑은 고딕', 9, 'bold')).pack(side='left')
        tk.Label(fe_top, text="비고", anchor='w', font=('맑은 고딕', 9, 'bold')).pack(side='left')
        self.rule_list = tk.Listbox(fe, height=3, font=('맑은 고딕', 9))
        self.rule_list.pack(fill='x', padx=8, pady=2)
        fe_in = tk.Frame(fe); fe_in.pack(fill='x', padx=8, pady=(2, 4))
        tk.Label(fe_in, text="이름:").pack(side='left')
        self._rule_name  = tk.StringVar()
        self._rule_time  = tk.StringVar()
        self._rule_note  = tk.StringVar()
        ttk.Entry(fe_in, textvariable=self._rule_name, width=10).pack(side='left', padx=(2, 8))
        tk.Label(fe_in, text="야근 시작:").pack(side='left')
        ttk.Entry(fe_in, textvariable=self._rule_time, width=7).pack(side='left', padx=(2, 4))
        tk.Label(fe_in, text="(예: 16:00)", font=('맑은 고딕', 8), fg='gray').pack(side='left', padx=(0, 8))
        tk.Label(fe_in, text="비고:").pack(side='left')
        ttk.Entry(fe_in, textvariable=self._rule_note, width=12).pack(side='left', padx=(2, 8))
        ttk.Button(fe_in, text="추가", command=self.add_employee_rule).pack(side='left', padx=2)
        ttk.Button(fe_in, text="선택 삭제", command=self.del_employee_rule).pack(side='left', padx=2)

        fs = ttk.LabelFrame(self.root, text="설정 — 수식 복원 템플릿 (저장됨)")
        fs.pack(fill='x', **pad)
        for label, var, key in [
            ("일반 양식 (설계·수주·안전·미래):", self.tmpl_normal, 'tmpl_normal'),
            ("시차출퇴근제 (서울설계):", self.tmpl_secha, 'tmpl_secha'),
        ]:
            row = tk.Frame(fs); row.pack(fill='x', padx=8, pady=2)
            tk.Label(row, text=label, width=26, anchor='w').pack(side='left')
            e = ttk.Entry(row, textvariable=var)
            e.pack(side='left', fill='x', expand=True, padx=(0, 6))
            self._register_drop_target(e, lambda ev, v=var, k=key: self._on_drop_tmpl(ev, v, k))
            ttk.Button(row, text="선택...",
                       command=lambda v=var, k=key: self.pick_tmpl(v, k)).pack(side='right')

        lf = ttk.LabelFrame(self.root, text="실행 로그")
        lf.pack(fill='both', expand=True, **pad)
        self.log = scrolledtext.ScrolledText(lf, height=10, font=('Consolas', 9), wrap='word')
        self.log.pack(fill='both', expand=True, padx=4, pady=4)

        # self.root를 실제 Toplevel/Tk 루트로 복원
        self.root = outer.winfo_toplevel()

    def _on_mode_change(self):
        mode = self.mode.get()
        self.f3a.pack_forget()
        self.f3b.pack_forget()
        self.f3c.pack_forget()

        if mode == 'clear':
            # 초기화는 출퇴근 기록부와 처리 기간이 필요 없다 — 화면에서 숨긴다.
            self._f1.pack_forget()
            self.f4.pack_forget()
            self.f3c.pack(fill='x', padx=12, pady=5, before=self.f5)
            self._f2.config(text="2. 초기화할 야근현황 파일  (.xlsx)  *필수")
            self.run_btn.config(text="▶  초 기 화  실 행", bg='#b45309')
            self._on_clr_change()
            return

        self._f1.pack(fill='x', padx=12, pady=5, before=self._f2)
        self.f4.pack(fill='x', padx=12, pady=5, before=self.f5)
        self.run_btn.config(text="▶  일 괄  실 행", bg='#2563eb')
        self.f5.config(text="5. 결과 저장 폴더")

        if mode == 'new':
            self.f3a.pack(fill='x', padx=12, pady=5, before=self.f4)
            self._f1.config(text="1. 출퇴근 기록부  (.xls / .xlsx)  *필수")
            self._f2.config(text="2. 전달 양식 파일  (.xlsx — 명단 추출 + 수식 소스)")
        else:
            self.f3b.pack(fill='x', padx=12, pady=5, before=self.f4)
            self._f2.config(text="2. 양식 파일  (.xlsx)")
            self._on_opt_change()

    def _on_clr_change(self):
        self.f5.config(
            text="5. 결과 저장 폴더  (덮어쓰기 선택됨 — 사용하지 않음)"
            if self.clr_overwrite.get() else "5. 결과 저장 폴더"
        )

    def _on_opt_change(self):
        need_att = self.opt_attendance.get()
        self._f1.config(
            text="1. 출퇴근 기록부  (.xls / .xlsx)  *필수" if need_att
            else "1. 출퇴근 기록부  (.xls / .xlsx)  (이 옵션 미선택 시 불필요)"
        )

    def _register_drop_target(self, widget, drop_handler):
        if not DND_AVAILABLE:
            return
        widget.drop_target_register(DND_FILES)
        widget.dnd_bind('<<Drop>>', drop_handler)
        widget.dnd_bind('<<DropEnter>>', lambda e: self._highlight(widget, True))
        widget.dnd_bind('<<DropLeave>>', lambda e: self._highlight(widget, False))

    def _highlight(self, widget, on):
        try:
            widget.config(bg=DROP_BG_HOVER if on else DROP_BG_NORMAL)
        except tk.TclError:
            pass

    def _on_drop_attendance(self, event):
        self._highlight(event.widget, False)
        files = expand_paths(parse_dropped_paths(event.data, self.root), ['.xls', '.xlsx'])
        for p in files:
            if p not in self.attendance_paths:
                self.attendance_paths.append(p)
                self.att_list.insert('end', os.path.basename(p))

    def _on_drop_forms(self, event):
        self._highlight(event.widget, False)
        files = expand_paths(parse_dropped_paths(event.data, self.root), ['.xlsx'])
        for p in files:
            if p not in self.form_paths:
                self.form_paths.append(p)
                self.form_list.insert('end', os.path.basename(p))
        if files and not self.output_dir.get():
            self.output_dir.set(os.path.dirname(files[0]))

    def _on_drop_tmpl(self, event, var, key):
        for p in parse_dropped_paths(event.data, self.root):
            p = p.strip().strip('{}').strip()
            if os.path.isfile(p) and p.lower().endswith('.xlsx'):
                var.set(p)
                _save_config({key: p})
                return

    def _on_drop_outdir(self, event):
        for p in parse_dropped_paths(event.data, self.root):
            p = p.strip().strip('{}').strip()
            d = p if os.path.isdir(p) else (os.path.dirname(p) if os.path.isfile(p) else None)
            if d:
                self.output_dir.set(d)
                _save_config({'output_dir': d})
                return

    def add_attendance(self):
        paths = filedialog.askopenfilenames(
            title="출퇴근 기록부 선택",
            filetypes=[("Excel", "*.xls *.xlsx"), ("All", "*.*")],
        )
        for p in paths:
            if p not in self.attendance_paths:
                self.attendance_paths.append(p)
                self.att_list.insert('end', os.path.basename(p))

    def clear_attendance(self):
        self.attendance_paths = []
        self.att_list.delete(0, 'end')

    def del_attendance(self):
        sel = self.att_list.curselection()
        if not sel:
            messagebox.showwarning("선택 필요", "삭제할 파일을 선택하세요."); return
        idx = sel[0]
        del self.attendance_paths[idx]
        self.att_list.delete(idx)

    def add_forms(self):
        self._add_forms_to_list(filedialog.askopenfilenames(
            title="양식 파일 선택", filetypes=[("Excel", "*.xlsx")]))

    def add_forms_folder(self):
        d = filedialog.askdirectory(title="양식 폴더 선택")
        if d:
            self._add_forms_to_list(sorted(
                str(p) for p in Path(d).glob("*.xlsx") if not p.name.startswith('~')))

    def _add_forms_to_list(self, paths):
        for p in paths:
            if p not in self.form_paths:
                self.form_paths.append(p)
                self.form_list.insert('end', os.path.basename(p))
            if not self.output_dir.get():
                d = os.path.dirname(p)
                self.output_dir.set(d)
                _save_config({'output_dir': d})

    def clear_forms(self):
        self.form_paths = []
        self.form_list.delete(0, 'end')

    def del_forms(self):
        sel = self.form_list.curselection()
        if not sel:
            messagebox.showwarning("선택 필요", "삭제할 파일을 선택하세요."); return
        idx = sel[0]
        del self.form_paths[idx]
        self.form_list.delete(idx)

    def _refresh_rule_list(self):
        self.rule_list.delete(0, 'end')
        for name, rule in self.employee_rules.items():
            t    = rule.get('야근시작', '')
            note = rule.get('비고', '')
            self.rule_list.insert('end', f"  {name:<10}  {t:<8}  {note}")

    def add_employee_rule(self):
        name = self._rule_name.get().strip().replace(' ', '')
        time_str = self._rule_time.get().strip()
        note = self._rule_note.get().strip()
        if not name:
            messagebox.showwarning("입력 필요", "이름을 입력하세요."); return
        import re as _re
        if not _re.match(r'^\d{1,2}:\d{2}$', time_str):
            messagebox.showwarning("형식 오류", "야근 시작 시간을 HH:MM 형식으로 입력하세요.\n예: 16:00"); return
        h, m = map(int, time_str.split(':'))
        if not (0 <= h <= 23 and 0 <= m <= 59):
            messagebox.showwarning("형식 오류", "유효하지 않은 시간입니다."); return
        self.employee_rules[name] = {'야근시작': f"{h:02d}:{m:02d}", '비고': note}
        _save_config({'employee_rules': self.employee_rules})
        self._refresh_rule_list()
        self._rule_name.set(''); self._rule_time.set(''); self._rule_note.set('')

    def del_employee_rule(self):
        sel = self.rule_list.curselection()
        if not sel:
            messagebox.showwarning("선택 필요", "삭제할 항목을 선택하세요."); return
        idx = sel[0]
        name = list(self.employee_rules.keys())[idx]
        del self.employee_rules[name]
        _save_config({'employee_rules': self.employee_rules})
        self._refresh_rule_list()

    def pick_tmpl(self, var, key):
        p = filedialog.askopenfilename(
            title="템플릿 파일 선택", filetypes=[("Excel", "*.xlsx")])
        if p:
            var.set(p)
            _save_config({key: p})

    def pick_output_dir(self):
        d = filedialog.askdirectory(title="저장 폴더 선택")
        if d:
            self.output_dir.set(d)
            _save_config({'output_dir': d})

    def _log(self, msg):
        self.log.insert('end', msg + '\n')
        self.log.see('end')
        self.root.update_idletasks()

    def clear_dates(self):
        self.user_start.set('')
        self.user_end.set('')

    def run(self):
        mode = self.mode.get()
        if not self.form_paths:
            messagebox.showwarning("입력 필요", "양식/대상 파일을 1개 이상 추가하세요.")
            return

        if mode == 'clear':
            self._run_clear()
            return

        if not self.output_dir.get():
            messagebox.showwarning("입력 필요", "저장 폴더를 선택하세요.")
            return

        try:
            user_start = parse_user_date(self.user_start.get())
            user_end   = parse_user_date(self.user_end.get())
        except ValueError as e:
            messagebox.showerror("날짜 형식 오류", str(e)); return
        if user_start and user_end and user_start > user_end:
            messagebox.showerror("날짜 오류", "시작일이 종료일보다 늦습니다."); return

        if mode == 'new':
            if not self.attendance_paths:
                messagebox.showwarning("입력 필요", "출퇴근 기록부를 추가하세요."); return
            try:
                ty = int(self.target_year.get())
                tm = int(self.target_month.get())
                if not (2000 <= ty <= 2100): raise ValueError("연도는 2000~2100")
                if not (1 <= tm <= 12):      raise ValueError("월은 1~12")
            except ValueError as e:
                messagebox.showerror("연월 오류", str(e)); return
            self.log.delete('1.0', 'end')
            self.run_btn.config(state='disabled', text="처리 중...")
            threading.Thread(
                target=self._worker_new,
                args=(self.attendance_paths[:], self.form_paths[:],
                      self.output_dir.get(), user_start, user_end, ty, tm),
                daemon=True).start()
        else:
            do_att  = self.opt_attendance.get()
            do_res  = self.opt_restore.get()
            do_pro  = self.opt_protect.get()
            if not do_att and not do_res:
                messagebox.showwarning("옵션 필요", "수정 옵션을 1개 이상 선택하세요."); return
            if do_att and not self.attendance_paths:
                messagebox.showwarning("입력 필요",
                    "'출퇴근 시간 추가/수정'을 선택했으나 출퇴근 기록부가 없습니다."); return
            self.log.delete('1.0', 'end')
            self.run_btn.config(state='disabled', text="처리 중...")
            threading.Thread(
                target=self._worker_modify,
                args=(self.attendance_paths[:], self.form_paths[:],
                      self.output_dir.get(), user_start, user_end,
                      do_att, do_res, do_pro,
                      self.tmpl_normal.get(), self.tmpl_secha.get(),
                      dict(self.employee_rules)),
                daemon=True).start()

    def _run_clear(self):
        overwrite = self.clr_overwrite.get()
        outdir = self.output_dir.get()

        if not overwrite and not outdir:
            messagebox.showwarning("입력 필요", "저장 폴더를 선택하거나 '원본 덮어쓰기'를 선택하세요.")
            return

        names = "\n".join(f"  · {os.path.basename(p)}" for p in self.form_paths[:10])
        more = f"\n  ... 외 {len(self.form_paths) - 10}개" if len(self.form_paths) > 10 else ""
        extra = "사유 텍스트까지 삭제합니다." if self.clr_reason.get() else "사유 텍스트는 보존합니다."

        if overwrite:
            msg = (f"아래 {len(self.form_paths)}개 파일의 출퇴근 데이터를 지우고\n"
                   f"원본 파일에 그대로 덮어씁니다. 되돌릴 수 없습니다.\n\n"
                   f"{names}{more}\n\n{extra}\n\n계속하시겠습니까?")
        else:
            msg = (f"아래 {len(self.form_paths)}개 파일의 출퇴근 데이터를 지워\n"
                   f"'_초기화' 사본으로 저장합니다.\n\n"
                   f"{names}{more}\n\n{extra}\n\n계속하시겠습니까?")

        if not messagebox.askyesno("출퇴근 데이터 초기화", msg, parent=self.root):
            return

        self.log.delete('1.0', 'end')
        self.run_btn.config(state='disabled', text="초기화 중...")
        threading.Thread(
            target=self._worker_clear,
            args=(self.form_paths[:], outdir, self.clr_reason.get(),
                  self.clr_restore.get(), overwrite,
                  dict(self.employee_rules)),
            daemon=True).start()

    def _worker_clear(self, forms, outdir, clear_reason, do_restore, overwrite,
                      employee_rules):
        try:
            self._log("[초기화 모드]")
            opts = ["수식 자동 복원" if do_restore else "수식 복원 안 함",
                    "사유 텍스트 삭제" if clear_reason else "사유 텍스트 보존",
                    "원본 덮어쓰기" if overwrite else "사본 저장"]
            self._log(f"옵션: {' / '.join(opts)}")
            self._log(f"\n[대상 파일 {len(forms)}개 처리...]")

            saved, failed = [], []
            for fp in forms:
                try:
                    r = process_clear_file(
                        fp, outdir, self._log,
                        clear_reason=clear_reason,
                        do_restore=do_restore,
                        overwrite=overwrite,
                        employee_rules=employee_rules)
                    if r:
                        saved.append(r)
                    else:
                        failed.append(os.path.basename(fp))
                except PermissionError:
                    self._log(f"    ❌ 파일이 열려 있어 저장할 수 없습니다 — Excel에서 닫고 다시 실행하세요")
                    failed.append(os.path.basename(fp))
                except Exception as e:
                    self._log(f"    ❌ 처리 실패: {e}")
                    failed.append(os.path.basename(fp))

            self._log(f"\n{'='*50}")
            self._log(f"✅ 초기화 완료 — 저장 {len(saved)}개, 실패 {len(failed)}개")
            for f in saved:
                self._log(f"   {os.path.basename(f)}")

            where = "원본 파일에 덮어썼습니다." if overwrite else f"저장 폴더:\n{outdir}"
            tail = f"\n\n실패 {len(failed)}개: {', '.join(failed)}" if failed else ""
            self.root.after(0, lambda: messagebox.showinfo(
                "완료", f"출퇴근 데이터 초기화가 끝났습니다.\n\n{where}{tail}"))
        except Exception as e:
            self._log(f"\n❌ 오류:\n{traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.root.after(0, lambda: self.run_btn.config(
                state='normal', text="▶  초 기 화  실 행"))

    def _worker_new(self, att, forms, outdir, user_start, user_end, ty, tm):
        try:
            process_all(att, forms, outdir, self._log, user_start, user_end, ty, tm,
                        self.employee_rules)
            self.root.after(0, lambda: messagebox.showinfo(
                "완료", f"새 양식 생성이 완료되었습니다.\n\n저장 폴더:\n{outdir}"))
        except Exception as e:
            self._log(f"\n❌ 오류:\n{traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.root.after(0, lambda: self.run_btn.config(
                state='normal', text="▶  일 괄  실 행"))

    def _worker_modify(self, att, forms, outdir, user_start, user_end,
                       do_att, do_res, do_pro, tmpl_normal, tmpl_secha,
                       employee_rules):
        try:
            self._log("[수정 모드]")
            opts = []
            if do_att: opts.append("출퇴근 추가/수정")
            if do_res: opts.append("수식 복원")
            if do_pro: opts.append("사유 보호")
            self._log(f"옵션: {' / '.join(opts)}")

            daily_map = {}
            if do_att and att:
                self._log(f"\n[출퇴근 파일 로딩] {len(att)}개...")
                daily_map, dups, total_rec, npeople = load_all_attendance(att, self._log)
                self._log(f"   유효 레코드 {total_rec:,}건, 총 인원 {npeople}명")
                if dups:
                    self._log(f"\n[⚠ 동명이인]")
                    for name, ids in dups.items():
                        self._log(f"   • '{name}' 사원번호 {ids}")

            self._log(f"\n[대상 파일 {len(forms)}개 처리...]")
            saved = []
            for fp in forms:
                try:
                    r = process_modify_file(
                        fp, daily_map, outdir, self._log,
                        do_att, do_res, do_pro,
                        tmpl_normal, tmpl_secha,
                        user_start, user_end,
                        employee_rules)
                    if r: saved.append(r)
                except Exception as e:
                    self._log(f"    ❌ 처리 실패: {e}")

            self._log(f"\n{'='*50}")
            self._log(f"✅ 완료 — 저장 파일 {len(saved)}개")
            for f in saved:
                self._log(f"   {os.path.basename(f)}")
            self.root.after(0, lambda: messagebox.showinfo(
                "완료", f"수정이 완료되었습니다.\n\n저장 폴더:\n{outdir}"))
        except Exception as e:
            self._log(f"\n❌ 오류:\n{traceback.format_exc()}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.root.after(0, lambda: self.run_btn.config(
                state='normal', text="▶  일 괄  실 행"))


def open_window(parent):
    win = tk.Toplevel(parent)
    OvertimeInputApp(win)
    return win
