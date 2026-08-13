"""
총무팀 자동화 허브 — 현장전산기기 신청내역 생성기
카카오톡 .txt + 기간 -> Excel 자동 생성
"""

import os
import re
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, scrolledtext

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from tkinterdnd2 import DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False


# ── 날짜 정규화 ───────────────────────────────────────────────────────
def normalize_date(s: str) -> str:
    digits = re.sub(r"\D", "", s.strip())
    if len(digits) == 8:
        y, m, d = digits[:4], digits[4:6], digits[6:8]
    elif len(digits) == 6:
        y, m, d = f"20{digits[:2]}", digits[2:4], digits[4:6]
    else:
        raise ValueError(f"날짜 형식을 인식할 수 없습니다: '{s}'")
    try:
        datetime(int(y), int(m), int(d))
    except ValueError:
        raise ValueError(f"올바른 날짜가 아닙니다: '{s}'")
    return f"{y}-{m}-{d}"


# ── Excel 스타일 상수 ─────────────────────────────────────────────────
HEADERS      = ["현장명", "지급일", "신청일", "품명", "수량", "비고"]
COL_WIDTHS   = [22, 12, 12, 38, 8, 36]
TITLE_FONT   = Font(name="맑은 고딕", bold=True, size=13)
HEADER_FONT  = Font(name="맑은 고딕", bold=True, size=10)
DATA_FONT    = Font(name="맑은 고딕", size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
HEADER_FILL  = PatternFill("solid", fgColor="D9E1F2")
SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")
_thin        = Side(style="thin", color="AAAAAA")
BORDER       = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# ── 키워드 목록 ───────────────────────────────────────────────────────
EQUIP_KW = [
    "노트북", "모니터", "복합기", "외장하드", "외장 하드",
    "NAS", "나스", "카트리지", "토너", "USB", "유에스비",
    "가방", "마우스", "키보드", "프린터", "잉크", "랜선",
    "공유기", "모뎀", "태블릿", "아이패드", "허브", "충전기",
    "어댑터", "SSD", "RAM", "메모리", "배터리", "베터리", "전원",
]
PAPER_KW = ["A4", "A3", "용지", "복사지"]

EXCLUDE_RE = re.compile(
    r"^(네|넵|알겠|감사|수고|확인했|완료|체크|OK\b|ok\b|ㅎ|ㄴ|ㅇ)"
    r"|사용.{0,5}현황|재고.{0,5}현황|점검.{0,5}완료|설치.{0,5}완료|반납"
    r"|\d+만\s*원|\d+만원|얼마죠|얼마에요|가격|정도요|정도 하고"
    r"|가져가셔야|가져가야|수리.{0,5}부탁",
    re.IGNORECASE,
)
QTY_RE = re.compile(
    r"(\d+)\s*(개|대|박스|권|장|롤|세트|set|EA|ea|TB|GB)",
    re.IGNORECASE,
)

_LINE_RE_OLD = re.compile(
    r"^\[(\d{4}-\d{2}-\d{2})\]\s+(?:오전|오후)\s+\d{1,2}:\d{2},\s+(.+?)\s*:\s+(.+)$"
)
_LINE_RE_NEW = re.compile(
    r"^\[(\d{4}-\d{2}-\d{2})\]\s+\[(.+?)\]\s+\[(?:오전|오후)\s+\d{1,2}:\d{2}\]\s+(.+)$"
)

def _match_line(line: str):
    return _LINE_RE_OLD.match(line) or _LINE_RE_NEW.match(line)


# ── 메시지 추출 ───────────────────────────────────────────────────────
def _parse_date_header(line: str):
    m = re.search(r"(\d{4})년\s+(\d{1,2})월\s+(\d{1,2})일", line)
    if m:
        return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()
    return None


def extract_messages(txt_path: str, start_str: str, end_str: str) -> list:
    start = datetime.strptime(start_str, "%Y-%m-%d").date()
    end   = datetime.strptime(end_str,   "%Y-%m-%d").date()
    lines = []
    current_date = None
    in_range = False

    with open(txt_path, encoding="utf-8-sig") as f:
        for line in f:
            line = line.rstrip("\n")
            d = _parse_date_header(line)
            if d is not None:
                current_date = d
                in_range = start <= d <= end
                continue
            if in_range and current_date and line.strip():
                lines.append(f"[{current_date}] {line}")
    return lines


# ── 규칙 기반 파서 ────────────────────────────────────────────────────
def _find_keyword(content: str, keywords: list) -> str:
    for kw in keywords:
        if kw.lower() in content.lower():
            return kw
    return None


def _extract_location(content: str, keyword: str, sender: str) -> str:
    idx = content.lower().find(keyword.lower())
    if idx > 0:
        before = content[:idx].strip()
        before = re.sub(r"[에서에의거기서 ]+$", "", before).strip()
        if before:
            return before
    after = content[idx + len(keyword):]
    site_m = re.search(r"([가-힣\w]+현장)", after)
    if site_m:
        return site_m.group(1)
    return f"({sender})"


def parse_messages(lines: list) -> dict:
    equipment = []
    paper     = []
    unclear   = []

    for line in lines:
        m = _match_line(line)
        if not m:
            continue
        date_str = m.group(1)
        sender   = m.group(2).split("(")[0].strip()
        content  = m.group(3).strip()

        if len(content) < 4 or EXCLUDE_RE.search(content):
            continue
        if re.match(r"^\[(사진|동영상|파일|이모티콘|삭제된)", content):
            continue

        is_paper = _find_keyword(content, PAPER_KW)
        is_equip = _find_keyword(content, EQUIP_KW)
        if not (is_paper or is_equip):
            continue

        qty_m    = QTY_RE.search(content)
        quantity = f"{qty_m.group(1)}{qty_m.group(2)}" if qty_m else "?"

        if is_paper:
            kw       = is_paper
            item     = f"{kw} 용지" if kw in ("A4", "A3") else kw
            location = _extract_location(content, kw, sender)
            paper.append({"현장명": location, "신청일": date_str,
                           "품명": item, "수량": quantity, "비고": ""})
        else:
            kw       = is_equip
            location = _extract_location(content, kw, sender)
            equipment.append({"현장명": location, "신청일": date_str,
                               "품명": kw, "수량": quantity, "비고": ""})

        if quantity == "?":
            unclear.append(f"  수량 불명확: [{date_str}] {content[:60]}")

    return {"equipment": equipment, "paper": paper, "unclear": unclear}


# ── Excel 작성 ────────────────────────────────────────────────────────
def _apply_section_header(ws, row: int, title: str) -> None:
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = Font(name="맑은 고딕", bold=True, size=11)
    cell.fill      = SECTION_FILL
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.row_dimensions[row].height = 18


def _apply_col_headers(ws, row: int) -> None:
    for col, (h, _) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[row].height = 16


def _write_rows(ws, entries: list, start_row: int) -> int:
    row = start_row
    for entry in entries:
        for col, key in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row, column=col, value=entry.get(key))
            cell.font      = DATA_FONT
            cell.border    = BORDER
            cell.alignment = CENTER if col in (2, 3, 5) else LEFT
        ws.row_dimensions[row].height = 15
        row += 1
    return row


def write_excel(data: dict, output_path: str, period: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "신청내역"

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1,
                         value=f"현장별 전산기기 지급 내역(진영) [{period}]")
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    current_row = 2
    equip = data.get("equipment", [])
    if equip:
        _apply_section_header(ws, current_row, "■ 전산기기")
        current_row += 1
        _apply_col_headers(ws, current_row)
        current_row += 1
        current_row = _write_rows(ws, equip, current_row)
        current_row += 1

    paper = data.get("paper", [])
    if paper:
        _apply_section_header(ws, current_row, "■ 용지")
        current_row += 1
        _apply_col_headers(ws, current_row)
        current_row += 1
        _write_rows(ws, paper, current_row)

    wb.save(output_path)


# ── GUI 클래스 (Toplevel 기반 — 허브 연동용) ─────────────────────────
class ITEquipApp:
    def __init__(self, parent):
        self.root = tk.Toplevel(parent)
        self.root.title("현장전산기기 신청내역 생성기")
        self.root.geometry("600x460")
        self.root.resizable(True, True)
        self.root.configure(bg="#f0f0f0")

        self._txt_path = tk.StringVar()
        self._start    = tk.StringVar(value="2026-01-01")
        self._end      = tk.StringVar(value=datetime.today().strftime("%Y-%m-%d"))
        self._out_dir  = tk.StringVar(value=os.path.expanduser("~\\Desktop"))

        self._build_ui()

    def _build_ui(self):
        p = {"padx": 10, "pady": 5}

        # ── 파일 (드래그앤드롭 지원) ──────────────────────────────────
        fr1 = tk.LabelFrame(
            self.root,
            text=" 카카오톡 대화파일 (.txt)  —  파일을 여기에 끌어다 놓거나 찾아보기 ",
            bg="#f0f0f0", padx=8, pady=4,
        )
        fr1.pack(fill="x", **p)

        self._txt_entry = tk.Entry(fr1, textvariable=self._txt_path)
        self._txt_entry.pack(side="left", fill="x", expand=True)
        tk.Button(fr1, text="찾아보기", command=self._browse_txt).pack(side="left", padx=(6, 0))

        if HAS_DND:
            try:
                self._txt_entry.drop_target_register(DND_FILES)
                self._txt_entry.dnd_bind("<<Drop>>", self._on_drop)
            except Exception:
                pass

        # ── 날짜 ──────────────────────────────────────────────────────
        fr2 = tk.LabelFrame(self.root, text=" 기간 (YYYY-MM-DD 또는 YYYYMMDD) ",
                            bg="#f0f0f0", padx=8, pady=4)
        fr2.pack(fill="x", **p)

        tk.Label(fr2, text="시작일", bg="#f0f0f0").pack(side="left")
        self._start_entry = tk.Entry(fr2, textvariable=self._start, width=13)
        self._start_entry.pack(side="left", padx=(4, 14))
        self._start_entry.bind("<FocusOut>", lambda _: self._fix_date(self._start))
        self._start_entry.bind("<Return>",   lambda _: self._fix_date(self._start))

        tk.Label(fr2, text="종료일", bg="#f0f0f0").pack(side="left")
        self._end_entry = tk.Entry(fr2, textvariable=self._end, width=13)
        self._end_entry.pack(side="left", padx=(4, 0))
        self._end_entry.bind("<FocusOut>", lambda _: self._fix_date(self._end))
        self._end_entry.bind("<Return>",   lambda _: self._fix_date(self._end))

        # ── 출력 폴더 ─────────────────────────────────────────────────
        fr3 = tk.LabelFrame(self.root, text=" 출력 폴더 ", bg="#f0f0f0", padx=8, pady=4)
        fr3.pack(fill="x", **p)
        tk.Entry(fr3, textvariable=self._out_dir).pack(side="left", fill="x", expand=True)
        tk.Button(fr3, text="찾아보기", command=self._browse_out).pack(side="left", padx=(6, 0))

        # ── 생성 버튼 ─────────────────────────────────────────────────
        self._btn = tk.Button(
            self.root, text="▶  신청내역 Excel 생성",
            command=self._on_generate,
            bg="#1976D2", fg="white",
            font=("맑은 고딕", 11, "bold"),
            padx=24, pady=8, relief="flat", cursor="hand2",
            activebackground="#1565C0", activeforeground="white",
        )
        self._btn.pack(pady=(6, 2))

        # ── 로그 ──────────────────────────────────────────────────────
        fr4 = tk.LabelFrame(self.root, text=" 진행 상황 ", bg="#f0f0f0", padx=8, pady=4)
        fr4.pack(fill="both", expand=True, **p)
        self._log = scrolledtext.ScrolledText(
            fr4, height=10, state="disabled",
            font=("맑은 고딕", 9), bg="#fafafa", relief="flat",
        )
        self._log.pack(fill="both", expand=True)

    def _on_drop(self, event):
        path = event.data.strip()
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        path = path.split("} {")[0].strip("{}")
        if path.lower().endswith(".txt"):
            self._txt_path.set(path)
        else:
            messagebox.showwarning("알림", ".txt 파일만 사용할 수 있습니다.", parent=self.root)

    def _fix_date(self, var: tk.StringVar):
        raw = var.get().strip()
        if not raw:
            return
        try:
            var.set(normalize_date(raw))
        except ValueError:
            pass

    def _browse_txt(self):
        path = filedialog.askopenfilename(
            parent=self.root,
            title="카카오톡 대화파일 선택",
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")],
        )
        if path:
            self._txt_path.set(path)

    def _browse_out(self):
        path = filedialog.askdirectory(parent=self.root, title="출력 폴더 선택")
        if path:
            self._out_dir.set(path)

    def _on_generate(self):
        if not self._txt_path.get():
            messagebox.showerror("오류", "카카오톡 대화파일을 선택해주세요.", parent=self.root)
            return
        try:
            start_normalized = normalize_date(self._start.get())
            end_normalized   = normalize_date(self._end.get())
            self._start.set(start_normalized)
            self._end.set(end_normalized)
        except ValueError as e:
            messagebox.showerror("날짜 오류", str(e), parent=self.root)
            return

        self._btn.config(state="disabled", text="생성 중...")
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            self._log_msg("1/3  메시지 추출 중...")
            lines = extract_messages(
                self._txt_path.get(),
                self._start.get(),
                self._end.get(),
            )
            if not lines:
                self._log_msg("⚠  해당 기간에 메시지가 없습니다.")
                self.root.after(0, lambda: messagebox.showwarning(
                    "알림", "해당 기간에 메시지가 없습니다.", parent=self.root))
                return
            self._log_msg(f"   → {len(lines)}행 추출 완료")

            self._log_msg("2/3  신청 항목 분류 중...")
            result  = parse_messages(lines)
            eq_cnt  = len(result.get("equipment", []))
            pa_cnt  = len(result.get("paper",     []))
            unclear = result.get("unclear", [])
            self._log_msg(f"   → 전산기기 {eq_cnt}건, 용지 {pa_cnt}건")
            if unclear:
                self._log_msg(f"   ⚠ 수량 불명확 {len(unclear)}건:")
                for u in unclear:
                    self._log_msg(u)

            start_s  = self._start.get().replace("-", "")
            end_s    = self._end.get().replace("-", "")
            fname    = f"현장전산기기_신청내역_{start_s}-{end_s}.xlsx"
            out_dir  = self._out_dir.get()
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir, fname)

            all_dates = sorted(
                d for d in (
                    [e.get("신청일", "") for e in result.get("equipment", [])]
                    + [e.get("신청일", "") for e in result.get("paper",    [])]
                ) if d
            )
            period = (
                f"{all_dates[0]} ~ {all_dates[-1]}" if all_dates
                else f"{self._start.get()} ~ {self._end.get()}"
            )

            self._log_msg(f"3/3  Excel 생성 중: {fname}")
            write_excel(result, out_path, period)
            self._log_msg(f"✔  저장 완료: {out_path}")

            note = "\n\n⚠ 수량 불명확 항목이 있습니다. Excel에서 확인해주세요." if unclear else ""
            self.root.after(0, lambda: messagebox.showinfo(
                "완료",
                f"전산기기: {eq_cnt}건\n용지: {pa_cnt}건\n\n{out_path}{note}",
                parent=self.root,
            ))

        except Exception as exc:
            self._log_msg(f"✖  오류: {exc}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(exc), parent=self.root))
        finally:
            self.root.after(0, lambda: self._btn.config(
                state="normal", text="▶  신청내역 Excel 생성"))

    def _log_msg(self, msg: str):
        self.root.after(0, self._append, msg)

    def _append(self, msg: str):
        self._log.config(state="normal")
        self._log.insert(tk.END, msg + "\n")
        self._log.see(tk.END)
        self._log.config(state="disabled")


# ── 허브 연동 인터페이스 ─────────────────────────────────────────────

def open_window(parent):
    app = ITEquipApp(parent)
    return app.root


# ── 단독 실행 ────────────────────────────────────────────────────────

if __name__ == "__main__":
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except ImportError:
        root = tk.Tk()
    root.withdraw()
    app = ITEquipApp(root)
    app.root.protocol("WM_DELETE_WINDOW", root.destroy)
    root.mainloop()
