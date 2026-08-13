"""
현장경비 입금확인 내역 조회 — 허브용 모듈
원본: app.py
변경: class App(tk.Tk) → class App(tk.Toplevel), open_window(parent) 함수 추가
"""

import sys
import os
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

SOURCE_FILE = ""
ALL_SHEETS = {}
SHEET_DATES = []
SHEET_DEPOSIT_DATES = {}

UNPAID_REASONS = {"유보", "유보금", "지급유보", "보류", "공사중지", "중지"}


def load_file(path):
    global SOURCE_FILE, ALL_SHEETS, SHEET_DATES, SHEET_DEPOSIT_DATES
    from openpyxl import load_workbook
    SOURCE_FILE = path
    ALL_SHEETS = pd.read_excel(path, sheet_name=None, header=None)
    SHEET_DATES = []
    SHEET_DEPOSIT_DATES = {}

    wb = load_workbook(path, data_only=True)
    for name in ALL_SHEETS:
        if '은행순' in name or name.startswith('Sheet'):
            continue
        try:
            dt = datetime.strptime(name.strip(), "%Y.%m")
            SHEET_DATES.append((dt, name))
        except ValueError:
            pass
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    if hasattr(cell.value, 'strftime'):
                        SHEET_DEPOSIT_DATES[name] = cell.value
    SHEET_DATES.sort()


def search_records(start_dt, end_dt, keyword):
    results = []
    for dt, sheet_name in SHEET_DATES:
        if not (start_dt <= dt <= end_dt):
            continue
        df = ALL_SHEETS[sheet_name]
        if df.empty or df.shape[0] < 2:
            continue
        deposit_dt = SHEET_DEPOSIT_DATES.get(sheet_name)
        deposit_str = deposit_dt.strftime("%Y-%m-%d") if deposit_dt else ""
        for _, row in df.iloc[1:].iterrows():
            site = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
            if not site or site == "nan":
                continue
            if keyword == "" or keyword in site:
                bigo = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ""
                if bigo == "nan":
                    bigo = ""
                results.append({
                    "입금일": deposit_str,
                    "현장명": site,
                    "은행명": str(row.iloc[2]) if pd.notna(row.iloc[2]) else "",
                    "지점명": str(row.iloc[3]) if pd.notna(row.iloc[3]) else "",
                    "계좌번호": str(row.iloc[4]) if pd.notna(row.iloc[4]) else "",
                    "예금주": str(row.iloc[5]) if pd.notna(row.iloc[5]) else "",
                    "금액": row.iloc[6] if pd.notna(row.iloc[6]) else 0,
                    "비고": bigo,
                    "미지급": bigo in UNPAID_REASONS,
                })
    return results


def export_excel(records, out_path, keyword, start_str, end_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "입금확인내역"

    cell_font = Font(name="맑은 고딕", size=11)
    bold_font = Font(name="맑은 고딕", bold=True, size=11)
    unpaid_font = Font(name="맑은 고딕", size=11, color="CC0000")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin")
    dbl = Side(style="double")
    border_data = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="92D050")
    fill_unpaid = PatternFill("solid", fgColor="FFE0E0")
    fill_sum    = PatternFill("solid", fgColor="D9E1F2")

    headers   = ["구분", "현장명", "은행명", "지점명", "계좌번호", "예금주", "금액", "입금일"]
    col_widths = [7,      38,       14,       12,       24,         10,       16,     14]
    h_border = Border(left=thin, right=thin, top=thin, bottom=dbl)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = cell_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = h_border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    total = 0
    for r_idx, rec in enumerate(records, 2):
        is_unpaid = rec["미지급"]
        fill = fill_unpaid if is_unpaid else PatternFill()
        font = unpaid_font if is_unpaid else cell_font

        vals = [
            r_idx - 1,
            rec["현장명"],
            rec["은행명"],
            rec["지점명"],
            rec["계좌번호"],
            rec["예금주"],
            rec["금액"],
            rec["입금일"],
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = font
            cell.alignment = center
            cell.border = border_data
            if fill.fill_type:
                cell.fill = fill
        ws.cell(row=r_idx, column=7).alignment = right
        if not is_unpaid:
            total += rec["금액"] if isinstance(rec["금액"], (int, float)) else 0
        ws.row_dimensions[r_idx].height = 18

    sum_row = len(records) + 2
    for r in range(2, sum_row):
        ws.cell(row=r, column=7).number_format = "#,##0"

    ws.merge_cells(f"A{sum_row}:F{sum_row}")
    sum_label = ws.cell(row=sum_row, column=1, value="합  계  (미지급 제외)")
    sum_label.font = bold_font
    sum_label.alignment = center
    sum_label.border = border_data
    sum_label.fill = fill_sum

    sum_val = ws.cell(row=sum_row, column=7, value=total)
    sum_val.font = bold_font
    sum_val.number_format = "#,##0"
    sum_val.alignment = right
    sum_val.border = border_data
    sum_val.fill = fill_sum

    for c in [8]:
        ws.cell(row=sum_row, column=c).border = border_data
        ws.cell(row=sum_row, column=c).fill = fill_sum

    paid_cnt   = sum(1 for r in records if not r["미지급"])
    unpaid_cnt = sum(1 for r in records if r["미지급"])
    note = ws.cell(row=sum_row + 1, column=1,
                   value=f"총 {len(records)}건  (지급 {paid_cnt}건 / 미지급 {unpaid_cnt}건)")
    note.font = Font(name="맑은 고딕", size=9, color="666666")

    cond = ws.cell(row=sum_row + 2, column=1,
                   value=f"조회기간: {start_str} ~ {end_str}" + (f"  /  현장: {keyword}" if keyword else ""))
    cond.font = Font(name="맑은 고딕", size=9, color="666666")

    wb.save(out_path)


class App(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("현장경비 입금확인 내역 조회")
        self.resizable(False, False)
        self.geometry("600x420")
        # 이 창만의 독립 상태 (전역 변수 문제 방지)
        self._source_file = ""
        self._all_sheets = {}
        self._sheet_dates = []
        self._sheet_deposit_dates = {}
        self._build_ui()

    def _load_file(self, path):
        from openpyxl import load_workbook
        self._source_file = path
        self._all_sheets = pd.read_excel(path, sheet_name=None, header=None)
        self._sheet_dates = []
        self._sheet_deposit_dates = {}

        wb = load_workbook(path, data_only=True)
        for name in self._all_sheets:
            if '은행순' in name or name.startswith('Sheet'):
                continue
            try:
                dt = datetime.strptime(name.strip(), "%Y.%m")
                self._sheet_dates.append((dt, name))
            except ValueError:
                pass
            if name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows():
                    for cell in row:
                        if hasattr(cell.value, 'strftime'):
                            self._sheet_deposit_dates[name] = cell.value
        self._sheet_dates.sort()

    def _search_records(self, start_dt, end_dt, keyword):
        results = []
        for dt, sheet_name in self._sheet_dates:
            if not (start_dt <= dt <= end_dt):
                continue
            df = self._all_sheets[sheet_name]
            if df.empty or df.shape[0] < 2:
                continue
            deposit_dt = self._sheet_deposit_dates.get(sheet_name)
            deposit_str = deposit_dt.strftime("%Y-%m-%d") if deposit_dt else ""
            for _, row in df.iloc[1:].iterrows():
                site = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                if not site or site == "nan":
                    continue
                if keyword == "" or keyword in site:
                    bigo = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else ""
                    if bigo == "nan":
                        bigo = ""
                    results.append({
                        "입금일": deposit_str,
                        "현장명": site,
                        "은행명": str(row.iloc[2]) if pd.notna(row.iloc[2]) else "",
                        "지점명": str(row.iloc[3]) if pd.notna(row.iloc[3]) else "",
                        "계좌번호": str(row.iloc[4]) if pd.notna(row.iloc[4]) else "",
                        "예금주": str(row.iloc[5]) if pd.notna(row.iloc[5]) else "",
                        "금액": row.iloc[6] if pd.notna(row.iloc[6]) else 0,
                        "비고": bigo,
                        "미지급": bigo in UNPAID_REASONS,
                    })
        return results

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        frm_file = tk.LabelFrame(self, text="원본 파일", font=("맑은 고딕", 10, "bold"))
        frm_file.pack(fill="x", **pad)

        self.file_var = tk.StringVar()
        tk.Entry(frm_file, textvariable=self.file_var, width=52, state="readonly").pack(side="left", padx=6, pady=6)
        tk.Button(frm_file, text="파일 열기", command=self._open_file, width=10).pack(side="left", padx=4)

        frm_cond = tk.LabelFrame(self, text="조회 조건", font=("맑은 고딕", 10, "bold"))
        frm_cond.pack(fill="x", **pad)

        tk.Label(frm_cond, text="기간 (시작)", font=("맑은 고딕", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=8)
        self.start_year = ttk.Combobox(frm_cond, width=7)
        self.start_year.grid(row=0, column=1, padx=2)
        tk.Label(frm_cond, text="년", font=("맑은 고딕", 10)).grid(row=0, column=2)
        self.start_month = ttk.Combobox(frm_cond, width=5, values=[str(m).zfill(2) for m in range(1, 13)])
        self.start_month.grid(row=0, column=3, padx=2)
        tk.Label(frm_cond, text="월", font=("맑은 고딕", 10)).grid(row=0, column=4)

        tk.Label(frm_cond, text="기간 (종료)", font=("맑은 고딕", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=8)
        self.end_year = ttk.Combobox(frm_cond, width=7)
        self.end_year.grid(row=1, column=1, padx=2)
        tk.Label(frm_cond, text="년", font=("맑은 고딕", 10)).grid(row=1, column=2)
        self.end_month = ttk.Combobox(frm_cond, width=5, values=[str(m).zfill(2) for m in range(1, 13)])
        self.end_month.grid(row=1, column=3, padx=2)
        tk.Label(frm_cond, text="월", font=("맑은 고딕", 10)).grid(row=1, column=4)

        tk.Label(frm_cond, text="현장명 (선택)", font=("맑은 고딕", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=8)
        self.keyword_var = tk.StringVar()
        tk.Entry(frm_cond, textvariable=self.keyword_var, width=30, font=("맑은 고딕", 10)).grid(row=2, column=1, columnspan=4, sticky="w", padx=2)
        tk.Label(frm_cond, text="(비워두면 전체)", font=("맑은 고딕", 9), fg="gray").grid(row=2, column=5, sticky="w")

        self.result_var = tk.StringVar(value="파일을 먼저 선택하세요.")
        tk.Label(self, textvariable=self.result_var, font=("맑은 고딕", 10), fg="#444").pack(pady=4)

        frm_btn = tk.Frame(self)
        frm_btn.pack(pady=10)
        tk.Button(frm_btn, text="조회 및 엑셀 저장", command=self._run,
                  font=("맑은 고딕", 11, "bold"), bg="#1F497D", fg="white",
                  width=18, height=2).pack(side="left", padx=8)
        tk.Button(frm_btn, text="닫기", command=self.destroy,
                  font=("맑은 고딕", 11), width=8, height=2).pack(side="left", padx=8)

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="현장경비 입금내역 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        self.result_var.set("파일 불러오는 중...")
        self.update()
        try:
            self._load_file(path)
            self.file_var.set(path)
            years = sorted(set(str(dt.year) for dt, _ in self._sheet_dates))
            self.start_year['values'] = years
            self.end_year['values'] = years
            if years:
                self.start_year.set(years[-1])
                self.end_year.set(years[-1])
                self.start_month.set("01")
                self.end_month.set("12")
            self.result_var.set(f"파일 로드 완료 — {len(self._sheet_dates)}개월 데이터")
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 실패:\n{e}")
            self.result_var.set("파일 읽기 실패")

    def _run(self):
        if not self._source_file:
            messagebox.showwarning("알림", "파일을 먼저 선택하세요.")
            return
        try:
            sy = int(self.start_year.get())
            sm = int(self.start_month.get())
            ey = int(self.end_year.get())
            em = int(self.end_month.get())
        except ValueError:
            messagebox.showwarning("알림", "연도와 월을 올바르게 입력하세요.")
            return

        start_dt = datetime(sy, sm, 1)
        end_dt = datetime(ey, em, 1)
        if start_dt > end_dt:
            messagebox.showwarning("알림", "시작 기간이 종료 기간보다 늦습니다.")
            return

        keyword = self.keyword_var.get().strip()
        records = self._search_records(start_dt, end_dt, keyword)

        if not records:
            messagebox.showinfo("결과 없음", "조건에 맞는 데이터가 없습니다.")
            self.result_var.set("검색 결과 없음")
            return

        start_str = f"{sy}.{sm:02d}"
        end_str = f"{ey}.{em:02d}"
        default_name = f"입금확인내역_{start_str}~{end_str}"
        if keyword:
            default_name += f"_{keyword}"
        default_name += ".xlsx"

        out_path = filedialog.asksaveasfilename(
            title="저장할 파일 선택",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 파일", "*.xlsx")]
        )
        if not out_path:
            return

        try:
            export_excel(records, out_path, keyword, start_str, end_str)
            paid = [r for r in records if not r["미지급"]]
            total = sum(r["금액"] for r in paid if isinstance(r["금액"], (int, float)))
            unpaid_cnt = len(records) - len(paid)
            self.result_var.set(f"저장 완료: 총 {len(records)}건 / 지급 {len(paid)}건 / 합계 {total:,.0f}원")
            messagebox.showinfo("완료",
                f"저장 완료!\n\n총 건수: {len(records)}건\n지급: {len(paid)}건 / 미지급: {unpaid_cnt}건\n합계(미지급 제외): {total:,.0f}원\n\n{out_path}")
            os.startfile(out_path)
        except Exception as e:
            messagebox.showerror("오류", f"저장 실패:\n{e}")


def open_window(parent):
    win = App(parent)
    return win
