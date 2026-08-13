import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import pytest

from mod_근로내용신고_core import (
    COL_DAY1,
    COL_JIKJONG,
    COL_JUMIN,
    COL_NAME,
    COL_NATION,
    COL_NTS_TAX,
    COL_NTS_YN,
    COL_PAY_MONTH,
    COL_RESIGN,
    COL_WAGE_TOTAL,
    COL_WORK_DAYS,
    DEFAULT_PARAMS,
    LedgerFormatError,
    build_reports,
    dedupe_people,
    detect_format,
    fill_missing_day_marks,
    group_by_site,
    is_foreigner,
    normalize_month,
    parse_ledger,
    report_filename,
    safe_filename,
    split_phone,
    summarize_sites,
)

TEMPLATE_HEADER = [
    "보험구분", "성명", "주민(외국인)등록번호", "국적코드", "체류자격코드",
    "전화(지역번호)", "전화(국번)", "전화(뒷번호)", "직종코드",
    *[f"{d}일" for d in range(1, 32)],
    "근로일수", "일평균근로시간", "보수지급기초일수", "보수총액", "임금총액",
    "이직사유코드", "보험료부과구분부호", "보험료부과구분사유",
    "국세청일용근로소득신고여부", "지급월", "총지급액", "비과세소득", "소득세", "지방소득세",
]


# --------------------------------------------------------------------- 픽스처

@pytest.fixture
def template(tmp_path):
    """전자신고용 양식의 '서식' 시트만 재현한 최소 템플릿."""
    path = tmp_path / "양식.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "서식"
    sheet.append(TEMPLATE_HEADER)
    workbook.save(path)
    return str(path)


def _write_format_a(path, site, people, month="2026년07월"):
    """[일용직 급여 지급명세서] 형식. people = [(코드, 성명, 주민번호, [근무일...], 급여, 전화)]"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(3, 9, "귀속:")
    sheet.cell(3, 11, month)
    sheet.cell(5, 23, "현장명 :")
    sheet.cell(5, 25, site)
    for col, label in ((1, "코드"), (2, "직종"), (3, "주민등록번호"), (4, "전화번호"), (5, "입사일")):
        sheet.cell(7, col, label)
    for col, label in ((1, "성명"), (2, "직급"), (3, "주 소")):
        sheet.cell(8, col, label)

    row = 9
    for code, name, jumin, days, pay, tel in people:
        sheet.cell(row, 1, code)
        sheet.cell(row, 3, jumin)
        if tel:
            sheet.cell(row, 4, tel)
        sheet.cell(row, 22, str(len(days)))
        sheet.cell(row, 23, len(days) * 8)
        sheet.cell(row, 25, pay)
        sheet.cell(row + 1, 1, name)
        for day in days:
            # 1~15일은 상단행 6~20열, 16~30일은 하단행 6~20열, 31일은 하단행 21열
            if day <= 15:
                sheet.cell(row, 5 + day, "●")
            elif day <= 30:
                sheet.cell(row + 1, 5 + (day - 15), "●")
            else:
                sheet.cell(row + 1, 21, "●")
        row += 2

    sheet.cell(row, 1, "총인원")
    workbook.save(path)
    return str(path)


def _write_format_b(path, site, people, month="2026-07"):
    """[현장별 일용직 급여내역] 형식. people = [(성명, 주민번호, 내/외, 근무일수, 급여)]"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ["현장코드", "현장명", "코드", "사원명", "주민(외국인)번호", "", "지급월", "귀속월",
              "최종근무일", "근무일", "근무시간", "지급액", "비과세", "소득세", "지방소득세",
              "임금총액", "공제총액", "차인지급액"]
    for col, label in enumerate(header, start=1):
        sheet.cell(5, col, label)

    row = 6
    for name, jumin, flag, days, pay in people:
        values = ["0074", site, "3300", name, flag, jumin, month, month,
                  "31", days, days * 8, pay, None, 1000, 100, pay, 0, pay]
        for col, value in enumerate(values, start=1):
            sheet.cell(row, col, value)
        row += 1

    sheet.cell(row, 1, "합  계")
    workbook.save(path)
    return str(path)


# ------------------------------------------------------------------ 셀 값 정리

def test_normalize_month_handles_known_formats():
    assert normalize_month("2026년07월") == "202607"
    assert normalize_month("2026-07") == "202607"
    assert normalize_month("2026-7") == "202607"
    assert normalize_month(None) is None
    assert normalize_month("알수없음") is None


def test_is_foreigner_detects_by_seventh_digit():
    assert is_foreigner("8804175340036") is True
    assert is_foreigner("8402151075736") is False


def test_is_foreigner_detects_by_ledger_flag():
    assert is_foreigner("8402151075736", "외") is True
    assert is_foreigner("8402151075736", "내") is False


def test_split_phone_splits_mobile_and_landline():
    assert split_phone("010-9207-7498") == ("010", "9207", "7498")
    assert split_phone("043-123-4567") == ("043", "123", "4567")


def test_split_phone_returns_blanks_when_unusable():
    assert split_phone(None) == (None, None, None)
    assert split_phone("1234") == (None, None, None)


def test_safe_filename_strips_path_characters():
    assert safe_filename("영동/종합:센터") == "영동_종합_센터"
    assert safe_filename("   ") == "현장미상"


def test_report_filename_falls_back_when_month_unknown():
    assert report_filename("영동", "202607") == "근로내용확인신고_영동_202607.xlsx"
    assert report_filename("영동", None) == "근로내용확인신고_영동_연월미상.xlsx"


# ------------------------------------------------------------------- 형식 판별

def test_detect_format_rejects_unknown_sheet():
    with pytest.raises(LedgerFormatError):
        detect_format([("아무", "관계없는", "표")])


def test_parse_ledger_detects_format_a(tmp_path):
    path = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [1, 15, 16, 31], 1200000, None)])
    fmt, people = parse_ledger(path)
    assert fmt == "A"
    assert len(people) == 1
    assert people[0]["site"] == "영동현장"
    assert people[0]["paymonth"] == "202607"


def test_parse_ledger_detects_format_b(tmp_path):
    path = _write_format_b(tmp_path / "b.xlsx", "영동현장",
                           [("서석재", "590820-1390519", "내", 4, 1200000)])
    fmt, people = parse_ledger(path)
    assert fmt == "B"
    assert people[0]["day_marks"] is None      # 형식 B에는 일자별 자료가 없다


# --------------------------------------------------------------- 일자 표시 위치

def test_format_a_maps_day_marks_to_correct_positions(tmp_path):
    path = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [1, 13, 16, 30, 31], 1200000, None)])
    _, people = parse_ledger(path)
    marks = people[0]["day_marks"]
    assert len(marks) == 31
    assert [i + 1 for i, m in enumerate(marks) if m] == [1, 13, 16, 30, 31]


def test_format_a_reads_declared_work_days(tmp_path):
    path = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [2, 3, 4], 600000, None)])
    _, people = parse_ledger(path)
    assert people[0]["days"] == 3


# ------------------------------------------------------------------- 현장별 분리

def test_group_by_site_splits_mixed_ledger(tmp_path):
    path = _write_format_b(tmp_path / "mixed.xlsx", "영동현장",
                           [("서석재", "590820-1390519", "내", 4, 1200000)])
    _, people = parse_ledger(path)
    people.append({**people[0], "site": "소수면현장", "name": "박미순"})
    sites = group_by_site(people)
    assert set(sites) == {"영동현장", "소수면현장"}


def test_group_by_site_labels_blank_site():
    assert "현장미상" in group_by_site([{"site": "", "name": "홍길동"}])


def test_build_reports_creates_one_file_per_site(tmp_path, template):
    a = _write_format_a(tmp_path / "영동.xlsx", "영동현장",
                        [("0800", "서석재", "590820-1390519", [1, 2], 400000, None)])
    b = _write_format_a(tmp_path / "소수면.xlsx", "소수면현장",
                        [("3200", "박미순", "620305-2389934", [5], 160000, None)])
    outdir = tmp_path / "out"

    results, loaded, duplicates = build_reports([a, b], template, str(outdir), DEFAULT_PARAMS)

    assert duplicates == 0
    assert len(loaded) == 2
    assert {r["site"] for r in results} == {"영동현장", "소수면현장"}
    assert sorted(os.listdir(outdir)) == [
        "근로내용확인신고_소수면현장_202607.xlsx",
        "근로내용확인신고_영동현장_202607.xlsx",
    ]


# --------------------------------------------------------------- 신고서 기입 내용

def _first_data_row(path):
    sheet = openpyxl.load_workbook(path, data_only=True)["서식"]
    return [cell.value for cell in sheet[2]]


def test_written_report_fills_core_fields(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1, 13, 31], 900000, "010-1234-5678")])
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), DEFAULT_PARAMS)
    row = _first_data_row(results[0]["path"])

    assert row[COL_NAME - 1] == "서석재"
    assert row[COL_JUMIN - 1] == "5908201390519"       # 하이픈 제거, 13자리
    assert row[COL_NATION - 1] == "100"                # 내국인
    assert row[COL_JIKJONG - 1] == "706"
    assert row[COL_WORK_DAYS - 1] == 3
    assert row[COL_WAGE_TOTAL - 1] == 900000
    assert row[COL_RESIGN - 1] == "1"                  # 보험구분 3 → 이직사유 필수


def test_written_report_marks_only_worked_days(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1, 13, 31], 900000, None)])
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), DEFAULT_PARAMS)
    row = _first_data_row(results[0]["path"])

    marked = [i + 1 for i in range(31) if row[COL_DAY1 - 1 + i]]
    assert marked == [1, 13, 31]


def test_foreigner_nationality_left_blank_for_manual_entry(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("3342", "윤드미트리", "880417-5340036", [1], 300000, None)])
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), DEFAULT_PARAMS)

    assert results[0]["foreigners"] == ["윤드미트리"]
    assert _first_data_row(results[0]["path"])[COL_NATION - 1] in (None, "")


def test_nts_columns_blank_when_not_reporting(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1], 300000, None)])
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), DEFAULT_PARAMS)
    row = _first_data_row(results[0]["path"])

    assert row[COL_NTS_YN - 1] is None
    assert row[COL_PAY_MONTH - 1] is None


def test_nts_columns_filled_when_reporting(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1], 300000, None)])
    params = {**DEFAULT_PARAMS, "nts": True}
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), params)
    row = _first_data_row(results[0]["path"])

    assert row[COL_NTS_YN - 1] == "Y"
    assert row[COL_PAY_MONTH - 1] == "202607"


def test_resign_code_omitted_for_industrial_accident_only(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1], 300000, None)])
    params = {**DEFAULT_PARAMS, "insurance": "1"}
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), params)

    assert _first_data_row(results[0]["path"])[COL_RESIGN - 1] is None


def test_blank_jikjong_is_left_empty(tmp_path, template):
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1], 300000, None)])
    params = {**DEFAULT_PARAMS, "jikjong": ""}
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), params)

    assert _first_data_row(results[0]["path"])[COL_JIKJONG - 1] is None


# ------------------------------------------------------- 일자 자료 없음 / 보완

def test_format_b_only_reports_missing_day_marks(tmp_path, template):
    ledger = _write_format_b(tmp_path / "b.xlsx", "영동현장",
                             [("서석재", "590820-1390519", "내", 3, 900000)])
    results, _, _ = build_reports([ledger], template, str(tmp_path / "out"), DEFAULT_PARAMS)

    assert results[0]["missing_days"] == ["서석재"]
    row = _first_data_row(results[0]["path"])
    assert all(row[COL_DAY1 - 1 + i] is None for i in range(31))
    assert row[COL_WORK_DAYS - 1] == 3      # 근로일수는 채워진다


def test_grid_file_fills_missing_day_marks(tmp_path, template):
    summary = _write_format_b(tmp_path / "b.xlsx", "영동현장",
                              [("서석재", "590820-1390519", "내", 3, 900000)])
    grid = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [4, 5, 6], 900000, None)])

    results, _, _ = build_reports([summary], template, str(tmp_path / "out"),
                                  DEFAULT_PARAMS, grid_paths=[grid])

    assert results[0]["missing_days"] == []
    row = _first_data_row(results[0]["path"])
    assert [i + 1 for i in range(31) if row[COL_DAY1 - 1 + i]] == [4, 5, 6]


# ------------------------------------------------------------------ 중복 병합

def test_same_person_in_both_formats_counted_once(tmp_path, template):
    """같은 현장 자료를 두 형식으로 함께 넣어도 인원이 두 배가 되지 않는다."""
    summary = _write_format_b(tmp_path / "b.xlsx", "영동현장",
                              [("서석재", "590820-1390519", "내", 3, 900000)])
    grid = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [4, 5, 6], 900000, None)])

    results, _, duplicates = build_reports([summary, grid], template,
                                           str(tmp_path / "out"), DEFAULT_PARAMS)

    assert duplicates == 1
    assert results[0]["count"] == 1
    assert results[0]["missing_days"] == []     # 일자표가 있는 쪽 값이 살아남는다
    row = _first_data_row(results[0]["path"])
    assert [i + 1 for i in range(31) if row[COL_DAY1 - 1 + i]] == [4, 5, 6]


def test_dedupe_keeps_day_marks_regardless_of_file_order(tmp_path, template):
    summary = _write_format_b(tmp_path / "b.xlsx", "영동현장",
                              [("서석재", "590820-1390519", "내", 3, 900000)])
    grid = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                           [("0800", "서석재", "590820-1390519", [4, 5, 6], 900000, None)])

    for order in ([summary, grid], [grid, summary]):
        results, _, _ = build_reports(order, template, str(tmp_path / "out"), DEFAULT_PARAMS)
        assert results[0]["count"] == 1
        assert results[0]["missing_days"] == []


def test_same_jumin_at_different_sites_kept_separately():
    people = [
        {"site": "영동현장", "jumin": "5908201390519", "name": "서석재", "day_marks": None},
        {"site": "소수면현장", "jumin": "5908201390519", "name": "서석재", "day_marks": None},
    ]
    deduped, duplicates = dedupe_people(people)
    assert duplicates == 0
    assert len(deduped) == 2


def test_fill_missing_day_marks_matches_on_jumin_only():
    people = [{"jumin": "5908201390519", "day_marks": None}]
    grid = [{"jumin": "9999999999999", "day_marks": [1] + [None] * 30}]
    assert fill_missing_day_marks(people, grid) == 0
    assert people[0]["day_marks"] is None


def test_summarize_sites_flags_day_count_mismatch():
    people = [{
        "site": "영동현장", "name": "서석재", "jumin": "5908201390519", "nat_flag": None,
        "days": 5, "wage": 900000, "paymonth": "202607",
        "day_marks": [1, 1, None] + [None] * 28,        # 2개만 표시 vs 근로일수 5
    }]
    assert summarize_sites(group_by_site(people))[0]["day_mismatch"] == ["서석재"]


def test_build_reports_rejects_empty_ledger(tmp_path, template):
    path = tmp_path / "empty.xlsx"
    _write_format_b(path, "영동현장", [])
    with pytest.raises(LedgerFormatError):
        build_reports([str(path)], template, str(tmp_path / "out"), DEFAULT_PARAMS)


def test_write_report_rejects_template_without_form_sheet(tmp_path):
    bad = tmp_path / "bad.xlsx"
    openpyxl.Workbook().save(bad)
    ledger = _write_format_a(tmp_path / "a.xlsx", "영동현장",
                             [("0800", "서석재", "590820-1390519", [1], 300000, None)])
    with pytest.raises(LedgerFormatError):
        build_reports([ledger], str(bad), str(tmp_path / "out"), DEFAULT_PARAMS)
